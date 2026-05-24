"""MIS Report endpoints — generates structured 8-period report with PDF/Excel downloads.

Period layout derived from ``as_of_date`` (e.g. 2026-05-13):
    P1: as_of_date .. as_of_date              (selected date / today)
    P2: first_of_current_month .. as_of_date  (MTD)
    P3..P7: previous 5 full months (newest first)
    P8: Total — computed as sum of P2..P7 for each row (no DB query)

Only metrics with known data sources are populated; others are placeholders
returning zero so the layout matches the source PDF and can be filled in later.
Known sources:
    - Electricity Units, DG Units, WIP, Dust to Boiler  → tbl_other_entries
    - Yarn Purchase / Sales (raw weights, by tran_type) → tbl_yarn_transaction + spinning_quality_mst
    - Raw Jute Arrival (sum of weight)                  → tbl_jute_received
    - Raw Jute Issued  (sum of net_wt)                  → assorting_entry
    - Raw Jute Closing = opening + received − issued    (opening = received − issued before period start)
    - Branded Bags / Bales (finished goods)             → tbl_daily_finishing.bales (×290 → Kgs)
    - Branded Bags / Bales (sales)                      → tbl_daily_finishing.issue_bales (×290 → Kgs)
    - Fabrics (in Kgs)                                  → tbl_daily_finishing.cuts × 30.2467
    - Pack Sheet (in Kgs)                               → tbl_daily_finishing.pack_sheet × 39
    - Mill Production (quality-wise, SUM net_weight)    → daily_doff_tbl + spinning_quality_mst
    - Hands                                             → daily_attendance: SUM(working_hours - idle_hours) / 8
"""

from __future__ import annotations

import calendar
import io
import traceback
from datetime import date, datetime, timedelta
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Request, Response
from fastapi.responses import StreamingResponse
from sqlalchemy import bindparam
from sqlalchemy.orm import Session
from sqlalchemy.sql import text

from src.authorization.utils import get_current_user_with_refresh
from src.config.db import get_tenant_db

router = APIRouter()

NUM_PERIODS = 8
TOTAL_COL_INDEX = NUM_PERIODS - 1  # last column is computed Total
TOTAL_SUM_START_INDEX = 1           # sum cols 2..(NUM_PERIODS-1) → store at TOTAL_COL_INDEX
TRAN_TYPE_PURCHASE = 1
TRAN_TYPE_SALES = 2


# ---------------------------------------------------------------------------
# Period helpers
# ---------------------------------------------------------------------------

def _last_day_of_month(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def _shift_month(d: date, delta_months: int) -> date:
    """Return first-of-month shifted by ``delta_months`` from ``d``."""
    total = d.year * 12 + (d.month - 1) + delta_months
    y, m = divmod(total, 12)
    return date(y, m + 1, 1)


def build_periods(as_of: date) -> list[dict[str, Any]]:
    """Return 8 period dicts.

    Layout: [selected date, MTD, prev-5-months..., Total].
    The Total period has no DB date range and is filled by ``_apply_totals``.
    """
    periods: list[dict[str, Any]] = []

    # P1: selected date (today)
    periods.append(
        {
            "label": as_of.strftime("%d-%m-%Y"),
            "from": as_of,
            "to": as_of,
            "sub": f"{as_of.strftime('%d-%m-%Y')} to {as_of.strftime('%d-%m-%Y')}",
            "is_total": False,
        }
    )

    # P2: 1st of current month .. as_of (MTD)
    p2_from = as_of.replace(day=1)
    periods.append(
        {
            "label": p2_from.strftime("%d-%m-%Y"),
            "from": p2_from,
            "to": as_of,
            "sub": f"{p2_from.strftime('%d-%m-%Y')} to {as_of.strftime('%d-%m-%Y')}",
            "is_total": False,
        }
    )

    # P3..P7: previous 5 full months
    for i in range(1, 6):
        first = _shift_month(as_of.replace(day=1), -i)
        last_day = _last_day_of_month(first.year, first.month)
        last = first.replace(day=last_day)
        periods.append(
            {
                "label": first.strftime("%d-%m-%Y"),
                "from": first,
                "to": last,
                "sub": f"{first.strftime('%d-%m-%Y')} to {last.strftime('%d-%m-%Y')}",
                "is_total": False,
            }
        )

    # P8: Total (computed)
    periods.append(
        {
            "label": "Total",
            "from": None,
            "to": None,
            "sub": "Total",
            "is_total": True,
        }
    )

    return periods


def _apply_totals(sections: list[dict[str, Any]]) -> None:
    """Set values[TOTAL_COL_INDEX] = sum(values[TOTAL_SUM_START_INDEX:TOTAL_COL_INDEX])."""
    for section in sections:
        for row in section["rows"]:
            row["values"][TOTAL_COL_INDEX] = sum(
                row["values"][TOTAL_SUM_START_INDEX:TOTAL_COL_INDEX]
            )


# ---------------------------------------------------------------------------
# Branch / company helpers
# ---------------------------------------------------------------------------

def _parse_branch_ids(raw: str | None) -> list[int]:
    if not raw:
        return []
    out: list[int] = []
    for part in raw.split(","):
        part = part.strip()
        if part.isdigit():
            out.append(int(part))
    return out


def _resolve_branch_ids(
    db: Session, co_id: int, branch_ids: list[int]
) -> list[int]:
    """If branch_ids empty, return all branches under co_id."""
    if branch_ids:
        return branch_ids
    rows = db.execute(
        text("SELECT branch_id FROM branch_mst WHERE co_id = :co_id"),
        {"co_id": co_id},
    ).fetchall()
    return [int(r[0]) for r in rows]


# ---------------------------------------------------------------------------
# Source queries (only for known tables)
# ---------------------------------------------------------------------------

def _other_entries_aggregates(
    db: Session, branch_ids: list[int], periods: list[dict[str, Any]]
) -> dict[str, list[float]]:
    """Sum elec_unit, dg_unit, wip_data, dust_boiler across each period."""
    keys = ("elec_unit", "dg_unit", "wip_data", "dust_boiler")
    result: dict[str, list[float]] = {k: [0.0] * NUM_PERIODS for k in keys}

    if not branch_ids:
        return result

    sql = text(
        """
        SELECT
            COALESCE(SUM(elec_unit), 0)   AS elec_unit,
            COALESCE(SUM(dg_unit), 0)     AS dg_unit,
            COALESCE(SUM(wip_data), 0)    AS wip_data,
            COALESCE(SUM(dust_boiler), 0) AS dust_boiler
        FROM tbl_other_entries
        WHERE branch_id IN :branch_ids
          AND tran_date BETWEEN :d_from AND :d_to
        """
    ).bindparams(bindparam("branch_ids", expanding=True))

    for idx, p in enumerate(periods):
        if p.get("is_total"):
            continue
        row = db.execute(
            sql,
            {"branch_ids": branch_ids, "d_from": p["from"], "d_to": p["to"]},
        ).fetchone()
        if row:
            m = row._mapping
            result["elec_unit"][idx] = float(m["elec_unit"] or 0)
            result["dg_unit"][idx] = float(m["dg_unit"] or 0)
            result["wip_data"][idx] = float(m["wip_data"] or 0)
            result["dust_boiler"][idx] = float(m["dust_boiler"] or 0)

    return result


def _jute_received_totals(
    db: Session, branch_ids: list[int], periods: list[dict[str, Any]]
) -> list[float]:
    """Sum tbl_jute_received.weight per period (Raw Jute Arrival in Kgs)."""
    result: list[float] = [0.0] * NUM_PERIODS
    if not branch_ids:
        return result

    sql = text(
        """
        SELECT COALESCE(SUM(weight), 0) AS total_weight
        FROM tbl_jute_received
        WHERE branch_id IN :branch_ids
          AND recv_date BETWEEN :d_from AND :d_to
        """
    ).bindparams(bindparam("branch_ids", expanding=True))

    for idx, p in enumerate(periods):
        if p.get("is_total"):
            continue
        row = db.execute(
            sql,
            {"branch_ids": branch_ids, "d_from": p["from"], "d_to": p["to"]},
        ).fetchone()
        if row:
            result[idx] = float(row._mapping["total_weight"] or 0)
    return result


def _jute_issued_totals(
    db: Session, branch_ids: list[int], periods: list[dict[str, Any]]
) -> list[float]:
    """Sum assorting_entry.net_wt per period (Raw Jute Issued in Kgs)."""
    result: list[float] = [0.0] * NUM_PERIODS
    if not branch_ids:
        return result

    sql = text(
        """
        SELECT COALESCE(SUM(net_wt), 0) AS total_weight
        FROM assorting_entry
        WHERE branch_id IN :branch_ids
          AND entry_date BETWEEN :d_from AND :d_to
        """
    ).bindparams(bindparam("branch_ids", expanding=True))

    for idx, p in enumerate(periods):
        if p.get("is_total"):
            continue
        row = db.execute(
            sql,
            {"branch_ids": branch_ids, "d_from": p["from"], "d_to": p["to"]},
        ).fetchone()
        if row:
            result[idx] = float(row._mapping["total_weight"] or 0)
    return result


def _hands_totals(
    db: Session, branch_ids: list[int], periods: list[dict[str, Any]]
) -> list[float]:
    """Hands per period = SUM(working_hours - idle_hours) / 8 from daily_attendance."""
    result: list[float] = [0.0] * NUM_PERIODS
    if not branch_ids:
        return result

    sql = text(
        """
        SELECT COALESCE(SUM(working_hours - COALESCE(idle_hours, 0)), 0) / 8 AS hands
        FROM daily_attendance
        WHERE branch_id IN :branch_ids
          AND COALESCE(is_active, 1) = 1
          AND attendance_date BETWEEN :d_from AND :d_to
        """
    ).bindparams(bindparam("branch_ids", expanding=True))

    for idx, p in enumerate(periods):
        if p.get("is_total"):
            continue
        row = db.execute(
            sql,
            {"branch_ids": branch_ids, "d_from": p["from"], "d_to": p["to"]},
        ).fetchone()
        if row:
            result[idx] = float(row._mapping["hands"] or 0)
    return result


def _mill_production_by_quality(
    db: Session, branch_ids: list[int], periods: list[dict[str, Any]]
) -> dict[str, list[float]]:
    """Return {quality_name: [period net_weight]} from daily_doff_tbl.

    Only qualities with at least one doff entry in the report periods appear.
    """
    out: dict[str, list[float]] = {}
    if not branch_ids:
        return out

    sql = text(
        """
        SELECT
            d.quality_id,
            sq.spg_quality,
            COALESCE(SUM(d.net_weight), 0) AS total_weight
        FROM daily_doff_tbl d
        LEFT JOIN spinning_quality_mst sq ON sq.spg_quality_mst_id = d.quality_id
        WHERE d.branch_id IN :branch_ids
          AND COALESCE(d.active, 1) = 1
          AND d.doff_date BETWEEN :d_from AND :d_to
        GROUP BY d.quality_id, sq.spg_quality
        """
    ).bindparams(bindparam("branch_ids", expanding=True))

    for idx, p in enumerate(periods):
        if p.get("is_total"):
            continue
        rows = db.execute(
            sql,
            {
                "branch_ids": branch_ids,
                "d_from": p["from"],
                "d_to": p["to"],
            },
        ).fetchall()
        for r in rows:
            m = r._mapping
            name = m["spg_quality"] or (
                f"Quality #{m['quality_id']}" if m["quality_id"] else "(unknown)"
            )
            out.setdefault(name, [0.0] * NUM_PERIODS)
            out[name][idx] = float(m["total_weight"] or 0)
    return out


def _daily_finishing_totals(
    db: Session, periods: list[dict[str, Any]]
) -> dict[str, list[float]]:
    """Sum tbl_daily_finishing columns (branding, bales, issue_bales) per period.

    Note: tbl_daily_finishing has no branch_id column, so this is not branch-filtered.
    Already tenant-scoped via the DB connection.
    """
    keys = ("branding", "bales", "issue_bales", "cuts", "pack_sheet")
    result: dict[str, list[float]] = {k: [0.0] * NUM_PERIODS for k in keys}

    sql = text(
        """
        SELECT
            COALESCE(SUM(branding),    0) AS branding,
            COALESCE(SUM(bales),       0) AS bales,
            COALESCE(SUM(issue_bales), 0) AS issue_bales,
            COALESCE(SUM(cuts),        0) AS cuts,
            COALESCE(SUM(pack_sheet),  0) AS pack_sheet
        FROM tbl_daily_finishing
        WHERE tran_date BETWEEN :d_from AND :d_to
        """
    )

    for idx, p in enumerate(periods):
        if p.get("is_total"):
            continue
        row = db.execute(
            sql, {"d_from": p["from"], "d_to": p["to"]}
        ).fetchone()
        if row:
            m = row._mapping
            result["branding"][idx] = float(m["branding"] or 0)
            result["bales"][idx] = float(m["bales"] or 0)
            result["issue_bales"][idx] = float(m["issue_bales"] or 0)
            result["cuts"][idx] = float(m["cuts"] or 0)
            result["pack_sheet"][idx] = float(m["pack_sheet"] or 0)

    return result


def _jute_opening_stock(
    db: Session, branch_ids: list[int], periods: list[dict[str, Any]]
) -> list[float]:
    """For each period, compute opening stock = received − issued before period start."""
    result: list[float] = [0.0] * NUM_PERIODS
    if not branch_ids:
        return result

    sql = text(
        """
        SELECT
            (SELECT COALESCE(SUM(weight), 0)
               FROM tbl_jute_received
              WHERE branch_id IN :branch_ids
                AND recv_date < :d_from) -
            (SELECT COALESCE(SUM(net_wt), 0)
               FROM assorting_entry
              WHERE branch_id IN :branch_ids
                AND entry_date < :d_from)
            AS opening
        """
    ).bindparams(bindparam("branch_ids", expanding=True))

    for idx, p in enumerate(periods):
        if p.get("is_total"):
            continue
        row = db.execute(
            sql,
            {"branch_ids": branch_ids, "d_from": p["from"]},
        ).fetchone()
        if row:
            result[idx] = float(row._mapping["opening"] or 0)
    return result


def _yarn_transactions_by_type_quality(
    db: Session,
    branch_ids: list[int],
    periods: list[dict[str, Any]],
    tran_type: int,
) -> dict[str, list[float]]:
    """Return {quality_name: [period weights]} for a given tran_type."""
    out: dict[str, list[float]] = {}
    if not branch_ids:
        return out

    sql = text(
        """
        SELECT
            COALESCE(sq.spg_quality, CONCAT('Quality #', y.quality_id)) AS quality_name,
            COALESCE(SUM(y.weight), 0) AS total_weight
        FROM tbl_yarn_transaction y
        LEFT JOIN spinning_quality_mst sq ON sq.spg_quality_mst_id = y.quality_id
        WHERE y.branch_id IN :branch_ids
          AND y.tran_type = :tran_type
          AND y.tran_date BETWEEN :d_from AND :d_to
        GROUP BY quality_name
        """
    ).bindparams(bindparam("branch_ids", expanding=True))

    for idx, p in enumerate(periods):
        if p.get("is_total"):
            continue
        rows = db.execute(
            sql,
            {
                "branch_ids": branch_ids,
                "tran_type": tran_type,
                "d_from": p["from"],
                "d_to": p["to"],
            },
        ).fetchall()
        for r in rows:
            m = r._mapping
            name = m["quality_name"] or "(unknown)"
            out.setdefault(name, [0.0] * NUM_PERIODS)
            out[name][idx] = float(m["total_weight"] or 0)
    return out


# ---------------------------------------------------------------------------
# Section builders
# ---------------------------------------------------------------------------

def _zeros() -> list[float]:
    return [0.0] * NUM_PERIODS


def _row(sl: int | str, label: str, values: list[float] | None = None) -> dict[str, Any]:
    return {
        "sl_no": sl,
        "label": label,
        "values": values if values is not None else _zeros(),
    }


def build_sections(
    db: Session,
    branch_ids: list[int],
    periods: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    other = _other_entries_aggregates(db, branch_ids, periods)
    yarn_purchases = _yarn_transactions_by_type_quality(
        db, branch_ids, periods, TRAN_TYPE_PURCHASE
    )
    yarn_sales = _yarn_transactions_by_type_quality(
        db, branch_ids, periods, TRAN_TYPE_SALES
    )
    raw_jute_arrival = _jute_received_totals(db, branch_ids, periods)
    raw_jute_issued = _jute_issued_totals(db, branch_ids, periods)
    raw_jute_opening = _jute_opening_stock(db, branch_ids, periods)
    raw_jute_closing = [
        raw_jute_opening[i] + raw_jute_arrival[i] - raw_jute_issued[i]
        for i in range(NUM_PERIODS)
    ]
    daily_finishing = _daily_finishing_totals(db, periods)
    # 290 kg per bale: Branded Bags (in Kgs) = bales * 290
    branded_bags_kgs = [v * 290 for v in daily_finishing["bales"]]
    # Sales: Branded Bags (in Kgs) = issue_bales * 290
    sales_branded_kgs = [v * 290 for v in daily_finishing["issue_bales"]]
    # Fabrics (in Kgs) = cuts * 30.2467
    fabrics_kgs = [v * 30.2467 for v in daily_finishing["cuts"]]
    # Pack Sheet (in Kgs) = pack_sheet * 39
    pack_sheet_kgs = [v * 39 for v in daily_finishing["pack_sheet"]]
    mill_production = _mill_production_by_quality(db, branch_ids, periods)
    hands_per_period = _hands_totals(db, branch_ids, periods)

    # Sum of yarn purchases (all qualities) - used for "Yarns/Fabrics Purchased"
    purchase_totals = _zeros()
    for vals in yarn_purchases.values():
        for i, v in enumerate(vals):
            purchase_totals[i] += v

    sections: list[dict[str, Any]] = []

    # 1. Raw Jute Report
    raw_jute_rows: list[dict[str, Any]] = [
        _row(1, "Raw Jute Arrival (in Kgs)", raw_jute_arrival),
        _row(2, "Yarns/Fabrics Purchased (in Kgs)", purchase_totals),
    ]
    # Quality-wise breakdown of yarn purchases — only qualities that actually
    # had a transaction in any of the report periods are listed.
    for quality_name in sorted(yarn_purchases.keys()):
        raw_jute_rows.append(
            _row("", f"    {quality_name} (in Kgs)", yarn_purchases[quality_name])
        )
    raw_jute_rows.extend(
        [
            _row(3, "Raw Jute Issued (in Kgs)", raw_jute_issued),
            _row(4, "Raw Jute Closing Stock (in Kgs)", raw_jute_closing),
        ]
    )
    sections.append({"title": "Raw Jute Report", "rows": raw_jute_rows})

    # 2. Finished Goods Stock
    sections.append(
        {
            "title": "Finished Goods Stock",
            "rows": [
                _row(1, "Branded Bags (in Kgs)", branded_bags_kgs),
                _row(2, "Branded Bags (in Bales)", daily_finishing["bales"]),
            ],
        }
    )

    # 3. Sales Summary
    sales_rows: list[dict[str, Any]] = [
        _row(1, "Branded Bags (in Kgs)", sales_branded_kgs),
        _row(2, "Branded Bags (in Bales)", daily_finishing["issue_bales"]),
    ]
    for i, quality_name in enumerate(sorted(yarn_sales.keys()), start=3):
        sales_rows.append(
            _row(i, f"{quality_name} (in Kgs)", yarn_sales[quality_name])
        )
    sections.append({"title": "Sales Summary", "rows": sales_rows})

    # 4. Stock Summary -- WIP + Dust real, others TODO
    sections.append(
        {
            "title": "Stock Summary",
            "rows": [
                _row(1, "Opening Stock"),
                _row(2, "Input"),
                _row(3, "Sales"),
                _row(4, "Jute Loss on Output 9%"),
                _row(5, "Closing Stock"),
                _row(6, "Ready Stock of Raw+Finished"),
                _row(7, "WIP", other["wip_data"]),
                _row(8, "Dust to Boiler", other["dust_boiler"]),
            ],
        }
    )

    # 5. Mill Production Summary -- dynamic per-quality from daily_doff_tbl
    mill_rows: list[dict[str, Any]] = []
    total_mill = _zeros()
    for i, quality_name in enumerate(sorted(mill_production.keys()), start=1):
        vals = mill_production[quality_name]
        for j, v in enumerate(vals):
            total_mill[j] += v
        mill_rows.append(_row(i, quality_name, vals))
    next_idx = len(mill_rows) + 1
    mill_rows.append(_row("", "Total Mill Shed Production", total_mill))
    mill_rows.append(_row(next_idx, "Fabrics", fabrics_kgs))
    mill_rows.append(_row(next_idx + 1, "Pack Sheet", pack_sheet_kgs))
    mill_rows.append(_row(next_idx + 2, "Stiched Bags"))
    mill_rows.append(_row(next_idx + 3, "Branded Bags", daily_finishing["branding"]))
    sections.append({"title": "Mill Production Summary", "rows": mill_rows})

    # 6. Analytical Report -- hands / elec / DG real, rest TODO
    sections.append(
        {
            "title": "Analytical Report",
            "rows": [
                _row(1, "Hands", hands_per_period),
                _row(2, "Hands per Ton"),
                _row(3, "Hands per Frame"),
                _row(4, "Wages"),
                _row(5, "Wages Per ton"),
                _row(6, "Electricity Units", other["elec_unit"]),
                _row(7, "DG Units", other["dg_unit"]),
                _row(8, "Units per ton"),
                _row(9, "No. of Frames run"),
                _row(10, "Per Frames Production"),
            ],
        }
    )

    # 7. Mill Production (New Shed) -- TODO: need shed split (mc_id → shed map)
    quality_labels = sorted(mill_production.keys())
    new_shed_rows: list[dict[str, Any]] = [
        _row(i + 1, lbl) for i, lbl in enumerate(quality_labels)
    ]
    new_shed_rows.append(_row("", "Total Mill Shed Production"))
    sections.append(
        {
            "title": "Mill Production Summary (in New Shed)",
            "rows": new_shed_rows,
            "page_break_before": True,
        }
    )

    # 8. Mill Production (Old Shed) -- TODO: need shed split (mc_id → shed map)
    old_shed_rows: list[dict[str, Any]] = [
        _row(i + 1, lbl) for i, lbl in enumerate(quality_labels)
    ]
    old_shed_rows.append(_row("", "Total Mill Shed Production"))
    sections.append(
        {"title": "Mill Production Summary (in Old Shed)", "rows": old_shed_rows}
    )

    # 9. Factory Production Summary -- branded bags & bales wired; rest TODO
    sections.append(
        {
            "title": "Factory Production Summary",
            "rows": [
                _row(1, "Fabrics (in Bags) (assumed 57/roll)"),
                _row(2, "Pack Sheet (in Bags)"),
                _row(3, "Cutting (in Bags)"),
                _row(4, "Heming (in Bags)"),
                _row(5, "Hiracle (in Bags)"),
                _row(6, "Branded Bags", daily_finishing["branding"]),
                _row(7, "Bales (of 500 Bags)", daily_finishing["bales"]),
            ],
        }
    )

    # 10. Target Production -- TODO config table
    sections.append(
        {
            "title": "Target Production",
            "rows": [
                _row(1, "Mill Production"),
                _row(2, "Factory Production"),
                _row(3, "Finishing Production"),
            ],
        }
    )

    _apply_totals(sections)
    return sections


# ---------------------------------------------------------------------------
# Main JSON endpoint
# ---------------------------------------------------------------------------

@router.get("/mis_report")
async def mis_report(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Return the MIS report as structured JSON."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        as_of_raw = request.query_params.get("as_of_date")
        if not as_of_raw:
            raise HTTPException(status_code=400, detail="as_of_date is required")
        try:
            as_of = datetime.strptime(as_of_raw, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(
                status_code=400, detail="as_of_date must be YYYY-MM-DD"
            )

        branch_ids = _parse_branch_ids(request.query_params.get("branch_id"))
        branch_ids = _resolve_branch_ids(db, int(co_id), branch_ids)

        # Company name for the report header
        co_row = db.execute(
            text("SELECT co_name FROM co_mst WHERE co_id = :co_id"),
            {"co_id": int(co_id)},
        ).fetchone()
        co_name = co_row[0] if co_row else "Company"

        periods = build_periods(as_of)
        sections = build_sections(db, branch_ids, periods)

        return {
            "company_name": co_name,
            "as_of_date": as_of.strftime("%Y-%m-%d"),
            "periods": [
                {"label": p["label"], "sub": p["sub"]} for p in periods
            ],
            "sections": sections,
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"mis_report error: {e}", flush=True)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Excel download endpoint
# ---------------------------------------------------------------------------

def _build_excel(report: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.pagebreak import Break

    wb = Workbook()
    ws = wb.active
    ws.title = "MIS Report"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="B7E1CD")
    section_fill = PatternFill("solid", fgColor="FCE5CD")

    # Title row
    ws.cell(row=1, column=1, value=report["company_name"]).font = bold
    ws.cell(row=2, column=1, value="MIS Report").font = bold
    ws.cell(row=3, column=1, value=f"As of {report['as_of_date']}")

    # Period header rows
    header_row = 5
    ws.cell(row=header_row, column=1, value="Sl.No").font = bold
    ws.cell(row=header_row, column=2, value="Particulars").font = bold
    for i, p in enumerate(report["periods"], start=3):
        c = ws.cell(row=header_row, column=i, value=p["sub"])
        c.font = bold
        c.alignment = center
        c.fill = header_fill

    r = header_row + 1
    for section in report["sections"]:
        if section.get("page_break_before") and r > header_row + 1:
            ws.row_breaks.append(Break(id=r - 1))
        ws.cell(row=r, column=1, value=section["title"]).font = bold
        ws.cell(row=r, column=1).fill = section_fill
        r += 1
        for row in section["rows"]:
            ws.cell(row=r, column=1, value=row["sl_no"])
            ws.cell(row=r, column=2, value=row["label"])
            for i, v in enumerate(row["values"], start=3):
                cell = ws.cell(row=r, column=i, value=v if v else None)
                cell.alignment = center
            r += 1

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 38
    for col in range(3, 3 + NUM_PERIODS):
        ws.column_dimensions[chr(64 + col)].width = 14

    # A4 landscape, fit to one page wide
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    # Repeat title + period header rows on every printed page
    ws.print_title_rows = f"1:{header_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/mis_report_excel")
async def mis_report_excel(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Download MIS report as .xlsx."""
    data = await mis_report(request, response, db, token_data)  # reuse logic
    xlsx = _build_excel(data)
    filename = f"mis_report_{data['as_of_date']}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# PDF download endpoint
# ---------------------------------------------------------------------------

def _build_pdf(report: dict[str, Any]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    left_margin = right_margin = 8 * mm
    page_w, page_h = landscape(A4)
    avail_w = page_w - left_margin - right_margin

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    flowables: list = []

    def add_heading():
        flowables.append(
            Paragraph(f"<b>{report['company_name']}</b>", styles["Title"])
        )
        flowables.append(
            Paragraph(
                f"MIS Report — As of {report['as_of_date']}", styles["Heading3"]
            )
        )
        flowables.append(Spacer(1, 6))

    add_heading()

    header = ["Sl.No", "Particulars"] + [p["sub"] for p in report["periods"]]

    # Split sections into groups separated by page breaks
    groups: list[list[dict[str, Any]]] = [[]]
    for section in report["sections"]:
        if section.get("page_break_before") and groups[-1]:
            groups.append([])
        groups[-1].append(section)

    num_period_cols = len(report["periods"])
    sl_w = 8 * mm
    particulars_w = 40 * mm
    period_w = (avail_w - sl_w - particulars_w) / num_period_cols
    col_widths = [sl_w, particulars_w] + [period_w] * num_period_cols

    def render_group(group_sections: list[dict[str, Any]]) -> Table:
        data_rows: list[list[Any]] = [header]
        section_row_indices: list[int] = []
        for section in group_sections:
            section_row_indices.append(len(data_rows))
            data_rows.append([section["title"]] + [""] * (len(header) - 1))
            for row in section["rows"]:
                vals = [
                    "" if not v else f"{v:,.2f}".rstrip("0").rstrip(".")
                    for v in row["values"]
                ]
                data_rows.append([row["sl_no"], row["label"], *vals])

        table = Table(data_rows, repeatRows=1, colWidths=col_widths)
        base_style: list = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#B7E1CD")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]
        for idx in section_row_indices:
            base_style.append(
                ("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#FCE5CD"))
            )
            base_style.append(("FONTNAME", (0, idx), (-1, idx), "Helvetica-Bold"))
            base_style.append(("SPAN", (1, idx), (-1, idx)))
        table.setStyle(TableStyle(base_style))
        return table

    for i, group in enumerate(groups):
        if not group:
            continue
        if i > 0:
            flowables.append(PageBreak())
            add_heading()
        flowables.append(render_group(group))

    doc.build(flowables)
    buf.seek(0)
    return buf.read()


@router.get("/mis_report_pdf")
async def mis_report_pdf(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Download MIS report as .pdf."""
    data = await mis_report(request, response, db, token_data)
    pdf = _build_pdf(data)
    filename = f"mis_report_{data['as_of_date']}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


__all__ = ["router", "build_periods", "build_sections"]


# Silence unused warning for the `date`/`timedelta` re-exports if linters complain
_ = (date, timedelta)
