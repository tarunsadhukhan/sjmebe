"""HRMS Bio Attendance bulk upload endpoints.

Flow:
  1. Client POSTs file to `/bio_att_upload`.
     - If `temp_bio_attendance_table` already has rows, the request is
       refused with "Another process Working".
     - Otherwise the file is parsed and bulk-inserted into the temp table
       *without any validation* (fast path).
     - A background job is queued that streams rows from the temp table,
       validates them against `bio_attendance_table` (duplicate check across
       the 10-column signature), inserts the survivors, and finally truncates
       the temp table.
  2. Client polls `/bio_att_excel_status/{job_id}` for progress.
  3. Client may call `/bio_att_clear` at any time to wipe the temp table.

Endpoints:
  - GET  /bio_att_list
  - POST /bio_att_upload
  - POST /bio_att_clear
  - GET  /bio_att_excel_status/{job_id}
"""

from __future__ import annotations

import csv
import io
import os
import traceback
import uuid
from datetime import datetime, date, time as dt_time, timedelta

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    HTTPException,
    Request,
    Response,
    UploadFile,
)
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.sql import text

from src.authorization.utils import get_current_user_with_refresh
from src.config.db import extract_subdomain_from_request, get_engine, get_tenant_db

router = APIRouter()

BATCH_SIZE = 2000  # rows per INSERT into the temp table


# ---------------------------------------------------------------------------
# Job-status persistence (bio_att_jobs table — survives across uvicorn workers)
# ---------------------------------------------------------------------------

_JOB_FIELDS = (
    "status", "total", "processed", "inserted",
    "duplicates", "invalid", "message", "error",
)

_JOB_UPSERT_SQL = text(
    """
    INSERT INTO bio_att_jobs
        (job_id, status, total, processed, inserted, duplicates, invalid, message, error)
    VALUES
        (:job_id, :status, :total, :processed, :inserted, :duplicates, :invalid, :message, :error)
    ON DUPLICATE KEY UPDATE
        status     = COALESCE(VALUES(status), status),
        total      = COALESCE(VALUES(total), total),
        processed  = COALESCE(VALUES(processed), processed),
        inserted   = COALESCE(VALUES(inserted), inserted),
        duplicates = COALESCE(VALUES(duplicates), duplicates),
        invalid    = COALESCE(VALUES(invalid), invalid),
        message    = COALESCE(VALUES(message), message),
        error      = COALESCE(VALUES(error), error)
    """
)

_JOB_SELECT_SQL = text(
    """
    SELECT job_id, status, total, processed, inserted, duplicates, invalid,
           message, error, created_at, updated_at
    FROM bio_att_jobs WHERE job_id = :job_id
    """
)


def _set_job(db: Session, job_id: str, **fields) -> None:
    """UPSERT a row in bio_att_jobs. Unspecified fields are left untouched."""
    params: dict = {f: None for f in _JOB_FIELDS}
    params.update({k: v for k, v in fields.items() if k in _JOB_FIELDS})
    params["job_id"] = job_id
    db.execute(_JOB_UPSERT_SQL, params)
    db.commit()


def _get_job(db: Session, job_id: str) -> dict | None:
    row = db.execute(_JOB_SELECT_SQL, {"job_id": job_id}).fetchone()
    if not row:
        return None
    d = dict(row._mapping)
    for ts_field in ("created_at", "updated_at"):
        v = d.get(ts_field)
        if isinstance(v, datetime):
            d[ts_field] = v.strftime("%Y-%m-%d %H:%M:%S")
    return d


# Columns that compose the duplicate signature in bio_attendance_table
DUP_COLUMNS = [
    "emp_code",
    "emp_anme",
    "bio_id",
    "log_date",
    "company_name",
    "department",
    "designation",
    "employement_type",
    "device_direction",
    "device_name",
]

# Mapping accepted column headers (lowercased) -> table column name.
# Both the device-style CSV and the table-style headers are accepted.
HEADER_ALIASES = {
    "employee code": "emp_code",
    "emp_code": "emp_code",

    "employee name": "emp_anme",
    "emp_anme": "emp_anme",
    "emp_name": "emp_anme",

    "employee code in device": "bio_id",
    "bio_id": "bio_id",

    "logdate": "log_date",
    "log date": "log_date",
    "log_date": "log_date",

    "company": "company_name",
    "company_name": "company_name",

    "department": "department",
    "designation": "designation",

    "employement type": "employement_type",
    "employment type": "employement_type",
    "employement_type": "employement_type",

    "direction": "device_direction",
    "device_direction": "device_direction",

    "device name": "device_name",
    "device_name": "device_name",
}

IGNORED_HEADERS = {"category"}


# ---------------------------------------------------------------------------
# Listing
# ---------------------------------------------------------------------------


@router.get("/bio_att_list")
async def bio_att_list(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Paginated bio-attendance listing."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        page = int(request.query_params.get("page", 1))
        limit = int(request.query_params.get("limit", 10))
        search = request.query_params.get("search")
        offset = max(page - 1, 0) * limit

        params: dict = {"limit": limit, "offset": offset}
        clauses: list[str] = []

        if search:
            clauses.append(
                "(emp_code LIKE :s OR emp_anme LIKE :s "
                " OR device_name LIKE :s OR department LIKE :s)"
            )
            params["s"] = f"%{search}%"

        # Per-column filters (whitelist -> safe column reference).
        # Frontend sends `f_<field>` matching DataGrid column field names.
        bio_filterable: dict[str, str] = {
            "emp_code":         "emp_code",
            "emp_anme":         "emp_anme",
            "bio_id":           "CAST(bio_id AS CHAR)",
            "log_date":         "CAST(log_date AS CHAR)",
            "department":       "department",
            "designation":      "designation",
            "device_direction": "device_direction",
            "device_name":      "device_name",
            "company_name":     "company_name",
            "employement_type": "employement_type",
        }
        for fld, expr in bio_filterable.items():
            v = request.query_params.get(f"f_{fld}")
            if v is None:
                continue
            v = v.strip()
            if not v:
                continue
            pname = f"f_{fld}"
            clauses.append(f"{expr} LIKE :{pname}")
            params[pname] = f"%{v}%"

        # Eb No (popup): exact emp_code match.
        emp_eq = (request.query_params.get("emp_code_eq") or "").strip()
        if emp_eq:
            clauses.append("emp_code = :emp_code_eq")
            params["emp_code_eq"] = emp_eq

        # Date-range filter: from_date / to_date applied to log_date.
        from_date = (request.query_params.get("from_date") or "").strip()
        to_date   = (request.query_params.get("to_date")   or "").strip()
        if from_date:
            clauses.append("DATE(log_date) >= :from_date")
            params["from_date"] = from_date
        if to_date:
            clauses.append("DATE(log_date) <= :to_date")
            params["to_date"] = to_date

        where = (" WHERE " + " AND ".join(clauses)) if clauses else ""

        rows = db.execute(
            text(f"""
                SELECT bio_att_id, emp_code, emp_anme, bio_id, log_date,
                       company_name, department, designation,
                       employement_type, device_direction, device_name
                FROM bio_attendance_table
                {where}
                ORDER BY log_date DESC, bio_att_id DESC
                LIMIT :limit OFFSET :offset
            """),
            params,
        ).fetchall()

        count_params = {k: v for k, v in params.items() if k not in ("limit", "offset")}
        total = db.execute(
            text(f"SELECT COUNT(*) AS cnt FROM bio_attendance_table {where}"),
            count_params,
        ).fetchone()

        data = []
        for r in rows:
            d = dict(r._mapping)
            ld = d.get("log_date")
            if isinstance(ld, datetime):
                d["log_date"] = ld.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(ld, date):
                d["log_date"] = ld.isoformat()
            data.append(d)

        return {
            "data": data,
            "total": int(total.cnt) if total else 0,
            "page": page,
            "limit": limit,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/daily_att_list")
async def daily_att_list(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Paginated daily_attendance_process_table listing."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        page = int(request.query_params.get("page", 1))
        limit = int(request.query_params.get("limit", 10))
        search = request.query_params.get("search")
        offset = max(page - 1, 0) * limit

        params: dict = {"limit": limit, "offset": offset}
        clauses: list[str] = []

        if search:
            clauses.append(
                "(o.emp_code LIKE :s "
                " OR CONCAT_WS(' ', p.first_name, IFNULL(p.middle_name,''), IFNULL(p.last_name,'')) LIKE :s "
                " OR b.department LIKE :s "
                " OR b.designation LIKE :s "
                " OR d.spell_name LIKE :s "
                " OR d.attendance_type LIKE :s "
                " OR d.device_name LIKE :s "
                " OR CAST(d.attendance_date AS CHAR) LIKE :s)"
            )
            params["s"] = f"%{search}%"

        # Per-column filters (whitelist mapped to qualified SQL expressions).
        daily_filterable: dict[str, str] = {
            "daily_att_proc_id": "CAST(d.daily_att_proc_id AS CHAR)",
            "bio_id":            "CAST(d.bio_id AS CHAR)",
            "emp_code":          "o.emp_code",
            "emp_name":          "TRIM(CONCAT_WS(' ', p.first_name, IFNULL(p.middle_name,''), IFNULL(p.last_name,'')))",
            "department":        "b.department",
            "designation":       "b.designation",
            "attendance_date":   "CAST(d.attendance_date AS CHAR)",
            "spell_name":        "d.spell_name",
            "attendance_type":   "d.attendance_type",
            "attendance_source": "d.attendance_source",
            "check_in":          "CAST(d.check_in AS CHAR)",
            "check_out":         "CAST(d.check_out AS CHAR)",
            "Time_duration":     "CAST(d.Time_duration AS CHAR)",
            "Working_hours":     "CAST(d.Working_hours AS CHAR)",
            "Ot_hours":          "CAST(d.Ot_hours AS CHAR)",
            "device_name":       "d.device_name",
        }
        for fld, expr in daily_filterable.items():
            v = request.query_params.get(f"f_{fld}")
            if v is None:
                continue
            v = v.strip()
            if not v:
                continue
            pname = f"f_{fld}"
            clauses.append(f"{expr} LIKE :{pname}")
            params[pname] = f"%{v}%"

        # Eb No (popup): exact emp_code match.
        emp_eq = (request.query_params.get("emp_code_eq") or "").strip()
        if emp_eq:
            clauses.append("o.emp_code = :emp_code_eq")
            params["emp_code_eq"] = emp_eq

        # Date-range filter: from_date / to_date applied to attendance_date.
        from_date = (request.query_params.get("from_date") or "").strip()
        to_date   = (request.query_params.get("to_date")   or "").strip()
        if from_date:
            clauses.append("d.attendance_date >= :from_date")
            params["from_date"] = from_date
        if to_date:
            clauses.append("d.attendance_date <= :to_date")
            params["to_date"] = to_date

        where = (" WHERE " + " AND ".join(clauses)) if clauses else ""

        rows = db.execute(
            text(f"""
                SELECT d.daily_att_proc_id, d.bio_id,
                       o.emp_code AS emp_code,
                       TRIM(CONCAT_WS(' ', p.first_name,
                                          IFNULL(p.middle_name,''),
                                          IFNULL(p.last_name,''))) AS emp_name,
                       b.department AS department,
                       b.designation AS designation,
                       d.attendance_date, d.spell_name, d.attendance_type,
                       d.attendance_source, d.check_in, d.check_out,
                       d.Time_duration, d.Working_hours, d.Ot_hours,
                       d.spell_start_time, d.spell_end_time, d.spell_hours,
                       d.processed, d.device_name
                FROM daily_attendance_process_table d
                LEFT JOIN hrms_ed_official_details o ON o.eb_id = d.eb_id
                LEFT JOIN hrms_ed_personal_details p ON p.eb_id = d.eb_id
                LEFT JOIN bio_attendance_table b     ON b.bio_att_id = d.bio_id
                {where}
                ORDER BY d.attendance_date DESC, d.daily_att_proc_id DESC
                LIMIT :limit OFFSET :offset
            """),
            params,
        ).fetchall()

        count_params = {k: v for k, v in params.items() if k not in ("limit", "offset")}
        total = db.execute(
            text(
                f"""SELECT COUNT(*) AS cnt
                    FROM daily_attendance_process_table d
                    LEFT JOIN hrms_ed_official_details o ON o.eb_id = d.eb_id
                    LEFT JOIN hrms_ed_personal_details p ON p.eb_id = d.eb_id
                    LEFT JOIN bio_attendance_table b     ON b.bio_att_id = d.bio_id
                    {where}"""
            ),
            count_params,
        ).fetchone()

        def _fmt(v):
            if isinstance(v, datetime):
                return v.strftime("%Y-%m-%d %H:%M:%S")
            if isinstance(v, date):
                return v.isoformat()
            if isinstance(v, timedelta):
                # MySQL TIME columns come back as timedelta in PyMySQL.
                total_seconds = int(v.total_seconds())
                h, rem = divmod(total_seconds, 3600)
                m, s = divmod(rem, 60)
                return f"{h:02d}:{m:02d}:{s:02d}"
            return v

        data = [{k: _fmt(v) for k, v in r._mapping.items()} for r in rows]

        return {
            "data": data,
            "total": int(total.cnt) if total else 0,
            "page": page,
            "limit": limit,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_log_date(value):
    """Parse a cell value into a datetime. Returns None if invalid."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()
    if not s:
        return None
    fmts = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y",
    )
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _to_int_or_none(value):
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _str_or_empty(value):
    return "" if value is None else str(value).strip()


def _resolve_header(raw):
    if raw is None:
        return None
    key = str(raw).strip().lower()
    if key in IGNORED_HEADERS:
        return None
    return HEADER_ALIASES.get(key)


def _read_upload(file_bytes: bytes, filename: str) -> list[dict]:
    """Return parsed rows from CSV / xlsx upload. Only structural validation
    (required columns must be present); no row-level validation is done here."""
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"openpyxl not installed: {e}")

    is_csv = (filename or "").lower().endswith(".csv")

    if is_csv:
        try:
            text_content = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text_content = file_bytes.decode("latin-1")
        reader = csv.reader(io.StringIO(text_content))
        wb = openpyxl.Workbook()
        ws = wb.active
        for csv_row in reader:
            ws.append(csv_row)
    else:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=False)
    try:
        header_cells = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="File is empty")

    col_idx: dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        db_col = _resolve_header(cell.value)
        if db_col and db_col not in col_idx:
            col_idx[db_col] = idx

    missing = [c for c in DUP_COLUMNS if c not in col_idx]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}",
        )

    rows: list[dict] = []
    for row_cells in rows_iter:
        values = [c.value for c in row_cells]
        if all((v is None or (isinstance(v, str) and not v.strip())) for v in values):
            continue

        def _val(name: str):
            i = col_idx[name]
            return values[i] if i < len(values) else None

        rows.append({
            "emp_code": _str_or_empty(_val("emp_code")) or None,
            "emp_anme": _str_or_empty(_val("emp_anme")) or None,
            "bio_id": _to_int_or_none(_val("bio_id")),
            "log_date": _parse_log_date(_val("log_date")),
            "company_name": _str_or_empty(_val("company_name")) or None,
            "department": _str_or_empty(_val("department")) or None,
            "designation": _str_or_empty(_val("designation")) or None,
            "employement_type": _str_or_empty(_val("employement_type")) or None,
            "device_direction": _str_or_empty(_val("device_direction")) or None,
            "device_name": _str_or_empty(_val("device_name")) or None,
        })
    return rows


# SQL-Server DeviceLogs export headers (output of the "SQL Create" query).
# Keys are header text lowercased with spaces/underscores stripped.
# EmployeeId and CompanyId are present in the file but not stored.
_DEVICELOGS_HEADERS = {
    "devicelogid": "bio_att_log_id",
    "deviceid": "device_id",
    "userid": "bio_id",
    "logdate": "log_date",
    "direction": "device_direction",
    "employeecode": "emp_code",
    "employeename": "emp_anme",
}


def _read_devicelogs_upload(file_bytes: bytes, filename: str) -> list[dict]:
    """Parse a SQL-Server DeviceLogs export (csv/xlsx) — the output of the
    "SQL Create" query — and map each row straight onto bio_attendance_table:

        DeviceLogId  -> bio_att_log_id      EmployeeName -> emp_anme
        EmployeeCode -> emp_code            UserId       -> bio_id
        LogDate      -> log_date            Direction    -> device_direction
        DeviceId     -> device_id

    EmployeeId / CompanyId are ignored. Rows with an unparseable LogDate or a
    missing DeviceLogId are skipped (no other row-level validation)."""
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"openpyxl not installed: {e}")

    is_csv = (filename or "").lower().endswith(".csv")
    if is_csv:
        try:
            text_content = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text_content = file_bytes.decode("latin-1")
        reader = csv.reader(io.StringIO(text_content))
        wb = openpyxl.Workbook()
        ws = wb.active
        for csv_row in reader:
            ws.append(csv_row)
    else:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=False)
    try:
        header_cells = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="File is empty")

    col_idx: dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        if cell.value is None:
            continue
        key = str(cell.value).strip().lower().replace("_", "").replace(" ", "")
        mapped = _DEVICELOGS_HEADERS.get(key)
        if mapped and mapped not in col_idx:
            col_idx[mapped] = idx

    missing = [
        c for c in ("bio_att_log_id", "log_date", "emp_code") if c not in col_idx
    ]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=(
                "Missing required columns: "
                + ", ".join(missing)
                + ". Expected a DeviceLogs export (DeviceLogId, UserId, "
                "LogDate, Direction, EmployeeCode, EmployeeName, DeviceId)."
            ),
        )

    rows: list[dict] = []
    for row_cells in rows_iter:
        values = [c.value for c in row_cells]
        if all((v is None or (isinstance(v, str) and not v.strip())) for v in values):
            continue

        def _val(name: str):
            i = col_idx.get(name)
            if i is None or i >= len(values):
                return None
            return values[i]

        log_id = _to_int_or_none(_val("bio_att_log_id"))
        log_date = _parse_log_date(_val("log_date"))
        if log_id is None or log_date is None:
            continue

        rows.append({
            "bio_att_log_id": log_id,
            "emp_code": _str_or_empty(_val("emp_code")) or None,
            "emp_anme": _str_or_empty(_val("emp_anme")) or None,
            "bio_id": _to_int_or_none(_val("bio_id")),
            "log_date": log_date,
            "device_direction": _str_or_empty(_val("device_direction")) or None,
            "device_id": _to_int_or_none(_val("device_id")),
        })
    return rows


# ---------------------------------------------------------------------------
# SQL constants
# ---------------------------------------------------------------------------

TEMP_INSERT_SQL = text(
    """
    INSERT INTO temp_bio_attendance_table
        (emp_code, emp_anme, bio_id, log_date, company_name,
         department, designation, employement_type,
         device_direction, device_name)
    VALUES
        (:emp_code, :emp_anme, :bio_id, :log_date, :company_name,
         :department, :designation, :employement_type,
         :device_direction, :device_name)
    """
)

TEMP_COUNT_SQL = text("SELECT COUNT(*) AS cnt FROM temp_bio_attendance_table")
TEMP_DELETE_SQL = text("DELETE FROM temp_bio_attendance_table")

# Only rows captured from these devices are eligible to move into
# bio_attendance_table. Anything else is treated as invalid/ignored.
ALLOWED_DEVICE_NAMES = ("AIFace(Mars) Out-F", "AIFace(Mars) In-F")

# Count rows that would be rejected as invalid (missing emp_code or log_date,
# or device_name not in the allow-list).
INVALID_COUNT_SQL = text(
    """
    SELECT COUNT(*) AS cnt
    FROM temp_bio_attendance_table
    WHERE emp_code IS NULL OR emp_code = '' OR log_date IS NULL
       OR device_name IS NULL
       OR device_name NOT IN ('AIFace(Mars) Out-F', 'AIFace(Mars) In-F')
    """
)

# Set-difference move: temp MINUS bio (NULL-safe on all 10 cols).
# Equivalent to:
#   SELECT * FROM temp
#   MINUS
#   SELECT * FROM bio
# but expressed with NOT EXISTS so MySQL can use indexes.
BULK_INSERT_SQL = text(
    """
    INSERT INTO bio_attendance_table
        (emp_code, emp_anme, bio_id, log_date, company_name,
         department, designation, employement_type,
         device_direction, device_name)
    SELECT t.emp_code, t.emp_anme, t.bio_id, t.log_date, t.company_name,
           t.department, t.designation, t.employement_type,
           t.device_direction, t.device_name
    FROM temp_bio_attendance_table t
    WHERE t.emp_code IS NOT NULL AND t.emp_code <> ''
      AND t.log_date IS NOT NULL
      AND t.device_name IN ('AIFace(Mars) Out-F', 'AIFace(Mars) In-F')
      AND NOT EXISTS (
          SELECT 1 FROM bio_attendance_table b
          WHERE b.emp_code = t.emp_code
            AND b.log_date = t.log_date
            AND b.emp_anme <=> t.emp_anme
            AND b.bio_id <=> t.bio_id
            AND b.company_name <=> t.company_name
            AND b.department <=> t.department
            AND b.designation <=> t.designation
            AND b.employement_type <=> t.employement_type
            AND b.device_direction <=> t.device_direction
            AND b.device_name <=> t.device_name
      )
    """
)


# ---------------------------------------------------------------------------
# Link-master back-fill (runs after BULK_INSERT_SQL).
#
# tbl_master_bio_link_mst columns (per user):
#   match_type   CHAR(1)        -- 'E', 'D', 'O', 'B'
#   bio_data     VARCHAR(...)   -- raw text as it appears in bio_attendance_table
#   master_id    INT            -- internal id from sjm masters
#
# Mapping:
#   E : bio_data = bio_attendance_table.emp_code     -> eb_id
#   D : bio_data = bio_attendance_table.department   -> dept_id
#   O : bio_data = bio_attendance_table.designation  -> desig_id
#   B : bio_data = bio_attendance_table.device_name  -> device_id
# ---------------------------------------------------------------------------

# (match_type, bio_attendance target column, bio_attendance source column)
_LINK_UPDATE_TEMPLATES = (
    ("E", "eb_id",     "emp_code"),
    ("D", "dept_id",   "department"),
    ("O", "desig_id",  "designation"),
    ("B", "device_id", "device_name"),
)


def _backfill_links(db: Session) -> dict:
    """Populate eb_id / dept_id / desig_id / device_id on bio_attendance_table
    using tbl_master_bio_link_mst (columns: match_type, bio_data, master_id).

    After the link-master pass, any rows still missing eb_id are resolved
    directly from the employee master (hrms_ed_official_details) on emp_code —
    the link master is often not populated for every employee.

    Returns rowcounts per match_type (plus ``E_hrms`` for the fallback).
    Each UPDATE is wrapped so a failure on one (e.g. missing column on
    bio_attendance_table) doesn't abort the others."""
    counts: dict = {"E": 0, "D": 0, "O": 0, "B": 0, "E_hrms": 0, "errors": []}
    for code, target_col, match_col in _LINK_UPDATE_TEMPLATES:
        sql = text(
            f"""
            UPDATE bio_attendance_table b
            JOIN tbl_master_bio_link_mst m
                ON m.match_type = :code
               AND m.bio_data = b.{match_col}
            SET b.{target_col} = m.master_id
            """
        )
        try:
            res = db.execute(sql, {"code": code})
            counts[code] = int(res.rowcount or 0)
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            counts["errors"].append(f"{code}: {type(e).__name__}: {e}")

    # Fallback: rows still missing eb_id after the 'E' link-master pass are
    # matched straight against the active employee master on emp_code.
    fallback_sql = text(
        """
        UPDATE bio_attendance_table b
        JOIN hrms_ed_official_details o
            ON o.emp_code = b.emp_code
           AND o.active = 1
        SET b.eb_id = o.eb_id
        WHERE b.eb_id IS NULL
        """
    )
    try:
        res = db.execute(fallback_sql)
        counts["E_hrms"] = int(res.rowcount or 0)
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        counts["errors"].append(f"E_hrms: {type(e).__name__}: {e}")
    return counts


# ---------------------------------------------------------------------------
# Background validate-and-move worker
# ---------------------------------------------------------------------------


def _make_session(subdomain: str) -> Session:
    tenant_url = (
        f"mysql+pymysql://{os.getenv('DATABASE_USER')}:{os.getenv('DATABASE_PASSWORD')}"
        f"@{os.getenv('DATABASE_HOST')}:{os.getenv('DATABASE_PORT')}/{subdomain}"
    )
    engine = get_engine(tenant_url)
    SessionTenant = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    return SessionTenant()


def _run_validate_and_move(job_id: str, subdomain: str) -> None:
    """Set-based move: ``temp MINUS bio`` -> ``bio_attendance_table``.
    The temp table is **not** deleted afterwards (use /bio_att_clear).
    Status is persisted to ``bio_att_jobs`` so any uvicorn worker can read it.
    """
    db = _make_session(subdomain)

    try:
        total = int((db.execute(TEMP_COUNT_SQL).fetchone() or {"cnt": 0}).cnt)
        _set_job(
            db, job_id,
            status="running", total=total, processed=0,
            inserted=0, duplicates=0, invalid=0,
        )

        invalid = int((db.execute(INVALID_COUNT_SQL).fetchone() or {"cnt": 0}).cnt)

        # Single statement: insert rows that exist in temp but not in bio.
        result = db.execute(BULK_INSERT_SQL)
        inserted = int(result.rowcount or 0)
        db.commit()

        duplicates = max(total - invalid - inserted, 0)

        # Back-fill eb_id / dept_id / desig_id / device_id from the
        # link master. Best-effort — failures don't abort the upload.
        try:
            link_counts = _backfill_links(db)
            db.commit()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
            link_counts = {"errors": ["backfill failed"]}

        # Wipe the staging table now that the move has been committed.
        try:
            db.execute(TEMP_DELETE_SQL)
            db.commit()
        except Exception:
            db.rollback()

        link_msg = (
            f" Linked: E={link_counts.get('E', 0)}"
            f" (+{link_counts.get('E_hrms', 0)} from emp master), "
            f"D={link_counts.get('D', 0)}, "
            f"O={link_counts.get('O', 0)}, B={link_counts.get('B', 0)}."
        )
        if link_counts.get("errors"):
            link_msg += f" Link errors: {'; '.join(link_counts['errors'])}"

        _set_job(
            db, job_id,
            status="completed",
            total=total,
            processed=total,
            inserted=inserted,
            duplicates=duplicates,
            invalid=invalid,
            message="Upload completed." + link_msg,
        )
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        try:
            _set_job(
                db, job_id,
                status="failed",
                error=f"{e}\n{traceback.format_exc()}"[:60000],
            )
        except Exception:
            pass
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.post("/bio_att_upload")
async def bio_att_upload(
    request: Request,
    response: Response,
    file: UploadFile = File(...),
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Parse a SQL-Server DeviceLogs export (the "SQL Create" query output)
    and insert every row straight into bio_attendance_table — no temp-table
    staging, no de-duplication. eb_id / dept_id / desig_id are then resolved
    with the same process the etrack flow uses (emp_code -> eb_id via the
    link master, dept/desig from daily_attendance with an hrms_ed_official
    fallback). Runs synchronously and returns the row count."""
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Empty file")

    rows = _read_devicelogs_upload(file_bytes, file.filename or "")
    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found")

    # Insert every parsed row directly into bio_attendance_table, in batches.
    try:
        inserted = 0
        for start in range(0, len(rows), BATCH_SIZE):
            batch = rows[start : start + BATCH_SIZE]
            res = db.execute(ETRACK_INSERT_SQL, batch)
            rc = int(res.rowcount or 0)
            inserted += rc if rc > 0 else len(batch)
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Insert into bio_attendance_table failed: {e}",
        )

    # Resolve eb_id / dept_id / desig_id — same process as the etrack flow.
    # Best-effort: a failure here doesn't undo the inserts above.
    resolve_msg = ""
    try:
        rr = _resolve_eb_id_from_link_master(db)
        db.commit()
        resolve_msg = (
            f" Resolved eb_id for {rr['resolved']} emp code(s), "
            f"{rr['updated']} row(s) updated "
            f"(dept/desig: {rr['from_daily_attendance']} from daily attendance, "
            f"{rr['fallback_official']} from emp master, "
            f"{rr['no_source']} with no source)."
        )
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        resolve_msg = f" (eb_id resolution failed: {e})"

    return {
        "message": (
            f"Inserted {inserted} row(s) into bio_attendance_table." + resolve_msg
        ),
        "inserted": inserted,
        "total": len(rows),
    }


@router.post("/bio_att_clear")
async def bio_att_clear(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Wipe the temp table. Manual override for stuck batches."""
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        existing = db.execute(TEMP_COUNT_SQL).fetchone()
        deleted = int(existing.cnt) if existing else 0
        db.execute(TEMP_DELETE_SQL)
        db.commit()
        return {"message": "Temp table cleared", "deleted": deleted}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/bio_att_excel_status/{job_id}")
async def bio_att_excel_status(
    job_id: str,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    job = _get_job(db, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return {"job_id": job_id, **job}


@router.get("/bio_att_temp_count")
async def bio_att_temp_count(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Return current row count of temp_bio_attendance_table."""
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")
    try:
        row = db.execute(TEMP_COUNT_SQL).fetchone()
        return {"count": int(row.cnt) if row else 0}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/bio_att_max_log_id")
async def bio_att_max_log_id(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Return MAX(bio_att_log_id) from bio_attendance_table for the given month.

    Query params: co_id (required), month=YYYY-MM (required). Returns
    {"max_id": 0} when the month has no rows.
    """
    co_id = request.query_params.get("co_id")
    month = request.query_params.get("month")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")
    if not month:
        raise HTTPException(status_code=400, detail="month (YYYY-MM) is required")
    try:
        y_str, m_str = month.split("-")
        year, mon = int(y_str), int(m_str)
    except (ValueError, AttributeError):
        raise HTTPException(status_code=400, detail="month must be in YYYY-MM format")
    try:
        row = db.execute(
            text(
                "SELECT MAX(bio_att_log_id) AS max_id "
                "FROM bio_attendance_table "
                "WHERE YEAR(log_date) = :y AND MONTH(log_date) = :m"
            ),
            {"y": year, "m": mon},
        ).mappings().first()
        max_id = int(row["max_id"]) if row and row["max_id"] is not None else 0
        print(f"Max bio_att_log_id for {month} is {max_id}", flush=True)
        return {"max_id": max_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# Manual bio-attendance entry  (manual_auto = 1)
#   Page: dashboardportal/hrms/bioAttendance
#
#   HR enters a single missed punch (employee + log_date + In/Out). Everything
#   else is resolved server-side so the manual row matches a device-sourced row:
#     bio_id          <- tbl_master_bio_link_mst (match_type='E', master_id=eb_id)
#     emp_code/anme   <- hrms_ed_official_details (+ personal details), active=1
#     dept_id/desig_id<- sub_dept_id / designation_id from official details
#     bio_att_log_id  <- global MAX(bio_att_log_id) + 1
#     device_id       <- 22 for 'in', else 14   (matches the etrack convention)
#     device_direction<- 'in' / 'out' (lowercased)
#     manual_auto     <- 1
# =============================================================================

@router.get("/bio_att_emp_search")
async def bio_att_emp_search(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Active-employee typeahead for the manual bio-attendance page.

    Query params: co_id (required), branch_id (required, csv allowed),
    search (optional). Returns up to 20 {eb_id, emp_code, emp_name}.
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")
    raw_branch_id = request.query_params.get("branch_id")
    if not raw_branch_id:
        raise HTTPException(status_code=400, detail="branch_id is required")
    try:
        branch_ids = [int(b) for b in raw_branch_id.split(",") if b.strip()]
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid branch_id format")
    if not branch_ids:
        raise HTTPException(status_code=400, detail="branch_id is required")

    search_raw = (request.query_params.get("search") or "").strip()
    params: dict = {"search": f"%{search_raw}%" if search_raw else None}
    branch_in = ", ".join(str(b) for b in branch_ids)  # validated ints — safe

    try:
        rows = db.execute(
            text(
                f"""
                SELECT o.eb_id,
                       o.emp_code,
                       TRIM(CONCAT_WS(' ', p.first_name,
                                           NULLIF(p.middle_name, ''),
                                           NULLIF(p.last_name, ''))) AS emp_name
                FROM hrms_ed_official_details o
                LEFT JOIN hrms_ed_personal_details p ON p.eb_id = o.eb_id
                WHERE o.active = 1
                  AND o.branch_id IN ({branch_in})
                  AND (:search IS NULL
                       OR o.emp_code LIKE :search
                       OR p.first_name LIKE :search
                       OR p.last_name LIKE :search
                       OR CONCAT_WS(' ', p.first_name, p.last_name) LIKE :search)
                ORDER BY o.emp_code
                LIMIT 20
                """
            ),
            params,
        ).fetchall()
        return {"data": [dict(r._mapping) for r in rows]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/bio_att_manual_list")
async def bio_att_manual_list(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Paginated listing of manually-entered punches (manual_auto = 1).

    Query params: co_id (required), branch_id (required, csv allowed),
    page, limit, search (emp_code / name / direction).
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")
    raw_branch_id = request.query_params.get("branch_id")
    if not raw_branch_id:
        raise HTTPException(status_code=400, detail="branch_id is required")
    try:
        branch_ids = [int(b) for b in raw_branch_id.split(",") if b.strip()]
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail="Invalid branch_id format")
    if not branch_ids:
        raise HTTPException(status_code=400, detail="branch_id is required")

    page = int(request.query_params.get("page", 1))
    limit = int(request.query_params.get("limit", 10))
    offset = max(page - 1, 0) * limit
    search = (request.query_params.get("search") or "").strip()
    branch_in = ", ".join(str(b) for b in branch_ids)  # validated ints — safe

    params: dict = {"limit": limit, "offset": offset}
    clauses = ["b.manual_auto = 1", f"o.branch_id IN ({branch_in})"]
    if search:
        clauses.append(
            "(b.emp_code LIKE :s OR b.emp_anme LIKE :s OR b.device_direction LIKE :s)"
        )
        params["s"] = f"%{search}%"
    where = " WHERE " + " AND ".join(clauses)

    try:
        rows = db.execute(
            text(
                f"""
                SELECT b.bio_att_id, b.bio_att_log_id, b.emp_code, b.emp_anme,
                       b.bio_id, b.log_date, b.device_direction, b.device_id,
                       b.eb_id, b.dept_id, b.desig_id
                FROM bio_attendance_table b
                JOIN hrms_ed_official_details o ON o.eb_id = b.eb_id
                {where}
                ORDER BY b.log_date DESC, b.bio_att_id DESC
                LIMIT :limit OFFSET :offset
                """
            ),
            params,
        ).fetchall()

        count_params = {k: v for k, v in params.items() if k not in ("limit", "offset")}
        total = db.execute(
            text(
                f"""
                SELECT COUNT(*) AS cnt
                FROM bio_attendance_table b
                JOIN hrms_ed_official_details o ON o.eb_id = b.eb_id
                {where}
                """
            ),
            count_params,
        ).fetchone()

        data = []
        for r in rows:
            d = dict(r._mapping)
            ld = d.get("log_date")
            if isinstance(ld, datetime):
                d["log_date"] = ld.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(ld, date):
                d["log_date"] = ld.isoformat()
            data.append(d)

        return {
            "data": data,
            "total": int(total.cnt) if total else 0,
            "page": page,
            "limit": limit,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/bio_att_manual_create")
async def bio_att_manual_create(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Insert a single manual punch into bio_attendance_table.

    Body: co_id (required), branch_id (required), eb_id (required),
          log_date (required, 'YYYY-MM-DD HH:MM[:SS]' or ISO), direction ('In'/'Out').

    All identity/bio fields are resolved server-side from the masters.
    """
    body = await request.json()
    co_id = body.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        eb_id = int(body.get("eb_id"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="eb_id is required")

    try:
        branch_id = int(body.get("branch_id"))
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="branch_id is required")

    direction_raw = (body.get("direction") or body.get("device_direction") or "").strip().lower()
    if direction_raw not in ("in", "out"):
        raise HTTPException(status_code=400, detail="direction must be 'In' or 'Out'")
    device_id = 22 if direction_raw == "in" else 14

    # Parse log_date — accept 'YYYY-MM-DDTHH:MM', 'YYYY-MM-DD HH:MM:SS', date-only.
    log_raw = (body.get("log_date") or "").strip()
    if not log_raw:
        raise HTTPException(status_code=400, detail="log_date is required")
    try:
        log_dt = datetime.fromisoformat(log_raw.replace("T", " "))
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail="log_date must be 'YYYY-MM-DD HH:MM:SS' or ISO format",
        )

    try:
        # Resolve employee identity (active, in this branch). Name/dept/desig
        # come from the employee master.
        emp = db.execute(
            text(
                """
                SELECT o.emp_code,
                       o.sub_dept_id   AS dept_id,
                       o.designation_id AS desig_id,
                       TRIM(CONCAT_WS(' ', p.first_name,
                                           NULLIF(p.middle_name, ''),
                                           NULLIF(p.last_name, ''))) AS emp_anme
                FROM hrms_ed_official_details o
                LEFT JOIN hrms_ed_personal_details p ON p.eb_id = o.eb_id
                WHERE o.eb_id = :eb_id AND o.active = 1 AND o.branch_id = :branch_id
                LIMIT 1
                """
            ),
            {"eb_id": eb_id, "branch_id": branch_id},
        ).mappings().first()
        if not emp:
            raise HTTPException(
                status_code=400,
                detail="Active employee not found for this eb_id / branch",
            )

        # emp_code (master_data) and bio_id (bio_dev_id) come from the 'E' link
        # master. Fall back to the employee-master emp_code if no link row exists.
        link = db.execute(
            text(
                "SELECT master_data, bio_dev_id FROM tbl_master_bio_link_mst "
                "WHERE match_type = 'E' AND master_id = :eb_id LIMIT 1"
            ),
            {"eb_id": eb_id},
        ).mappings().first()
        bio_id = link["bio_dev_id"] if link and link["bio_dev_id"] is not None else None
        emp_code = (link["master_data"] if link and link["master_data"] else None) or emp["emp_code"]

        params = {
            "emp_code": emp_code,
            "emp_anme": emp["emp_anme"],
            "bio_id": bio_id,
            "log_date": log_dt,
            "device_direction": direction_raw,
            "eb_id": eb_id,
            "dept_id": emp["dept_id"],
            "desig_id": emp["desig_id"],
            "device_id": device_id,
        }

        # bio_att_log_id = global MAX + 1, computed in a derived table so MySQL
        # allows referencing the target table inside INSERT ... SELECT.
        db.execute(
            text(
                """
                INSERT INTO bio_attendance_table
                    (bio_att_log_id, emp_code, emp_anme, bio_id, log_date,
                     device_direction, eb_id, dept_id, desig_id, device_id,
                     manual_auto)
                SELECT m.next_id, :emp_code, :emp_anme, :bio_id, :log_date,
                       :device_direction, :eb_id, :dept_id, :desig_id, :device_id,
                       1
                FROM (SELECT COALESCE(MAX(bio_att_log_id), 0) + 1 AS next_id
                      FROM bio_attendance_table) AS m
                """
            ),
            params,
        )
        db.commit()

        return {
            "message": "Manual bio-attendance entry saved",
            "bio_id_resolved": bio_id is not None,
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# Bio-attendance -> daily_attendance_process_table  (Process button)
# =============================================================================
#
# Algorithm (per emp x date):
#   * Validate emp_code against hrms_ed_official_details (active=1) -> eb_id.
#       Unmatched emp_codes for the date are returned as an XLSX download
#       and the entire process is aborted.
#   * Spell A:  IN-time window 05:30..09:59:59,  work window 06:00..14:00
#   * Spell B:  IN-time window 13:30..15:59:59,  work window 14:00..22:00
#       (No Spell C / night shift.)
#   * Working day vs weekly off:
#       - tbl_offday_mst.day_off stores int 0..6 (Sun..Sat).
#       - If the attendance_date weekday matches a day_off row -> off day.
#       - Off day spell row: attendance_type = 'O'
#       - Working day spell row: attendance_type = 'R'
#   * Overtime split:
#       - Only on working days.
#       - If raw duration > 8.5h, additionally insert one row with
#         spell_name = same A/B, attendance_type = 'O',
#         Working_hours = duration - 8.0,  Ot_hours = duration - 8.0.
#   * Re-run on same date = DELETE existing rows for that date, re-insert.

SPELL_A_IN_FROM = "05:30:00"
SPELL_A_IN_TO   = "10:00:00"  # exclusive
SPELL_A_START   = "06:00:00"
SPELL_A_END     = "14:00:00"

SPELL_B_IN_FROM = "13:30:00"
SPELL_B_IN_TO   = "16:00:00"  # exclusive
SPELL_B_START   = "14:00:00"
SPELL_B_END     = "22:00:00"

SPELL_HOURS = 8.0
OT_THRESHOLD_HOURS = 8.5


def _log_sql(label: str, stmt, params: dict | None = None) -> None:
    """Print a SQL statement + bind params to stdout (uvicorn console).
    Used for live debugging the Process endpoint."""
    try:
        sql_str = str(getattr(stmt, "text", stmt))
    except Exception:
        sql_str = repr(stmt)
    print(f"\n[bio_att_process] === {label} ===", flush=True)
    print(sql_str.strip(), flush=True)
    if params is not None:
        print(f"[bio_att_process] params = {params}", flush=True)


# 1. Find rows on :tran_date that are NOT fully linked
#    (any of eb_id / dept_id / desig_id / device_id is NULL).
#    These are reported as the "unmatched" xlsx and skipped during processing.
UNMATCHED_EMP_CODES_SQL = text(
    """
    SELECT b.emp_code, b.emp_anme, b.bio_id, b.log_date,
           b.company_name, b.department, b.designation,
           b.employement_type, b.device_direction, b.device_name
    FROM bio_attendance_table b
    WHERE DATE(b.log_date) = :tran_date
      AND (b.eb_id IS NULL
        OR b.dept_id IS NULL
        OR b.desig_id IS NULL
        OR b.device_id IS NULL)
    ORDER BY b.emp_code, b.log_date
    """
)

# 2. Did weekday match a day_off row? (Returns count > 0 if it's a weekly-off day.)
#    Convention: tbl_offday_mst.off_day uses 1=Sun, 2=Mon, ..., 7=Sat.
#    MySQL DAYOFWEEK() also returns 1=Sun..7=Sat, so use it directly.
IS_OFF_DAY_SQL = text(
    "SELECT COUNT(*) AS cnt FROM tbl_offday_mst WHERE off_day = DAYOFWEEK(:tran_date)"
)

# 3. Wipe re-run rows for the date.
DELETE_DAY_ROWS_SQL = text(
    "DELETE FROM daily_attendance_process_table WHERE attendance_date = :tran_date"
)

# 4. Fetch all punches for the spell window. We then pair them in Python
#    (1st=IN, 2nd=OUT, 3rd=IN, 4th=OUT, ...) and only count pairs whose
#    duration is at least 1 hour as Working_hours. check_in is the first
#    punch, check_out is the last punch.
#
#    Only rows where ALL four link ids (eb_id, dept_id, desig_id, device_id)
#    are populated are considered. Window: punches between :in_from and
#    :spell_end (inclusive of in_from, inclusive of spell_end+OT slack).
FETCH_SPELL_PUNCHES_SQL = text(
    """
    SELECT b.eb_id,
           b.bio_att_log_id,
           b.dept_id,
           b.desig_id,
           TIME(b.log_date)   AS punch_time,
           b.log_date         AS log_date
    FROM bio_attendance_table b
    WHERE DATE(b.log_date) = :tran_date
      AND b.eb_id     IS NOT NULL
      AND b.dept_id   IS NOT NULL
      AND b.desig_id  IS NOT NULL
      AND b.device_id IS NOT NULL
      AND TIME(b.log_date) >= :in_from
      AND TIME(b.log_date) <= :window_end
    ORDER BY b.eb_id, b.log_date
    """
)

# Insert one daily row per employee using bound params.
INSERT_SPELL_ROW_SQL = text(
    """
    INSERT INTO daily_attendance_process_table
        (eb_id, bio_id, dept_id, desig_id,
         attendance_date, spell_name, attendance_type,
         attendance_source, check_in, check_out,
         Time_duration, Working_hours, Ot_hours,
         spell_start_time, spell_end_time, spell_hours, processed)
    VALUES
        (:eb_id, :bio_id, :dept_id, :desig_id,
         :tran_date, :spell_name, :attendance_type,
         'BIO', :check_in, :check_out,
         :time_duration, :working_hours, :ot_hours,
         :spell_start, :spell_end, :spell_hours, 1)
    """
)

# Minimum pair duration that counts toward Working_hours.
MIN_PAIR_SECONDS = 3600  # 1 hour


def _bucket_hours(h: float) -> float:
    """Bucket hours: >=7 -> 8, [3, 7) -> 4, <3 -> 0."""
    if h >= 7:
        return 8.0
    if h >= 3:
        return 4.0
    return 0.0


def _parse_hms(s: str) -> dt_time:
    """Parse 'HH:MM:SS' into datetime.time."""
    h, m, sec = (int(x) for x in s.split(":"))
    return dt_time(h, m, sec)


def _to_seconds(t) -> int:
    """Convert a datetime.time / timedelta-ish value to seconds-of-day."""
    if isinstance(t, timedelta):
        return int(t.total_seconds())
    if isinstance(t, dt_time):
        return t.hour * 3600 + t.minute * 60 + t.second
    # Fallback: parse string 'HH:MM:SS'
    return _to_seconds(_parse_hms(str(t)))


def _process_one_spell(
    db: Session,
    *,
    tran_date: str,
    spell_name: str,
    in_from: str,
    in_to: str,
    spell_start: str,
    spell_end: str,
    is_off_day: bool,
) -> tuple[int, int]:
    """Build daily_attendance_process_table rows for this spell.

    Logic:
      * Pull every punch (per emp) in [in_from .. spell_end+slack] for the date.
      * The employee belongs to this spell iff their first punch is within
        [in_from .. in_to).
      * Total duration = (last_punch - first_punch). Skip if < 1 hour.
      * Working_hours_raw = min(duration, 8); ot_hours_raw = max(0, duration - 8).
      * Bucket each: >=7 -> 8, [3, 7) -> 4, <3 -> 0.
      * Time_duration = working_hours + ot_hours (post-bucket).
      * On an off day, ot_hours = 0.

    Returns (regular_inserted, 0). OT does not create a separate row.
    """
    in_from_sec = _to_seconds(_parse_hms(in_from))
    in_to_sec   = _to_seconds(_parse_hms(in_to))

    # Slack so that a late check-out after spell_end is still captured.
    # Use a wide window (effectively end-of-day) so OT punches several
    # hours past spell_end aren't dropped. The spell that an employee
    # belongs to is decided below by their FIRST punch falling within
    # [in_from .. in_to), so widening the upper bound is safe.
    OT_SLACK_HOURS = 16
    spell_end_t = _parse_hms(spell_end)
    window_end_dt = (datetime.combine(date.min, spell_end_t)
                     + timedelta(hours=OT_SLACK_HOURS))
    # cap at 23:59:59 (we don't cross midnight here)
    if window_end_dt.day != date.min.day:
        window_end = "23:59:59"
    else:
        window_end = window_end_dt.time().strftime("%H:%M:%S")

    fetch_params = {
        "tran_date": tran_date,
        "in_from": in_from,
        "window_end": window_end,
    }
    _log_sql(f"FETCH_SPELL_PUNCHES_SQL [spell={spell_name}]",
             FETCH_SPELL_PUNCHES_SQL, fetch_params)
    rows = db.execute(FETCH_SPELL_PUNCHES_SQL, fetch_params).fetchall()
    print(f"[bio_att_process]   -> fetched {len(rows)} punch row(s)", flush=True)

    # Group punches per employee, preserving SQL ORDER BY (eb_id, log_date).
    # Each tuple: (punch_time, bio_att_log_id, dept_id, desig_id, log_date).
    by_emp: dict[int, list[tuple]] = {}
    for r in rows:
        m = r._mapping
        by_emp.setdefault(m["eb_id"], []).append(
            (m["punch_time"], m["bio_att_log_id"], m["dept_id"], m["desig_id"], m["log_date"])
        )

    inserted_reg = 0
    inserted_ot  = 0
    base_attendance_type = "O" if is_off_day else "P"

    for eb_id, punches in by_emp.items():
        if not punches:
            continue
        first_time, first_bio, first_dept, first_desig, first_log_date = punches[0]
        first_sec = _to_seconds(first_time)
        # Employee only belongs to this spell if first punch is inside IN window.
        if not (in_from_sec <= first_sec < in_to_sec):
            continue

        last_time, _, _, _, last_log_date = punches[-1]
        last_sec = _to_seconds(last_time)

        # Total span: last punch - first punch (covers etrack data where
        # only IN/OUT are recorded, and gives the "full presence" duration).
        span_secs = max(0, last_sec - first_sec)
        if span_secs < MIN_PAIR_SECONDS:
            # Less than 1 hour on site -- skip row.
            continue

        total_hours    = round(span_secs / 3600.0, 2)
        # Working hours = duration capped at 8; OT = anything beyond 8.
        capped_working = round(min(SPELL_HOURS, total_hours), 2)
        if is_off_day:
            ot_raw = 0.0
        else:
            ot_raw = round(max(total_hours - SPELL_HOURS, 0.0), 2)
        # Bucket both: >=7 -> 8, [3, 7) -> 4, <3 -> 0.
        capped_working = _bucket_hours(capped_working)
        ot_hours = _bucket_hours(ot_raw)
        # Time_duration = total of bucketed working + OT.
        time_duration = round(capped_working + ot_hours, 2)

        reg_params = {
            "eb_id": int(eb_id),
            "bio_id": int(first_bio) if first_bio is not None else None,
            "dept_id": int(first_dept) if first_dept is not None else None,
            "desig_id": int(first_desig) if first_desig is not None else None,
            "tran_date": tran_date,
            "spell_name": spell_name,
            "attendance_type": base_attendance_type,
            "check_in": first_log_date,
            "check_out": last_log_date,
            "time_duration": time_duration,
            "working_hours": capped_working,
            "ot_hours": ot_hours,
            "spell_start": spell_start,
            "spell_end": spell_end,
            "spell_hours": SPELL_HOURS,
        }
        _log_sql(f"INSERT_SPELL_ROW_SQL [spell={spell_name} eb_id={eb_id}]",
                 INSERT_SPELL_ROW_SQL, reg_params)
        db.execute(INSERT_SPELL_ROW_SQL, reg_params)
        inserted_reg += 1

    inserted_ot = 0  # OT no longer creates a separate row.
    print(f"[bio_att_process]   -> rows inserted = {inserted_reg}", flush=True)
    return inserted_reg, inserted_ot


# ─────────────────────────────────────────────────────────────────────────────
# Etrack-specific processing. Used by /bio_att_etrack_process.
# May produce ONE OR TWO daily_attendance_process_table rows per (eb_id, date)
# depending on the first-entry window and last-entry time.
#
# Notation (decimal hours-of-day):
#   first_h = first_punch hour-of-day      e.g. 5:30 -> 5.5
#   last_h  = last_punch  hour-of-day      e.g. 17:00 -> 17.0
#   span    = last_h - first_h             (must be >= 1 hour to qualify)
#
# Rule 1 — first_h in [5, 9):
#   eff = last_h - 6
#   if eff >= 7        -> rec1 {spell:A,  type:R, working:8}
#   elif eff in [3,7)  -> rec1 {spell:A1, type:R, working:4}
#   O = eff - working_hours_of_rec1
#   if O >= 7          -> rec2 {spell:B,  type:O, ot:8}
#   elif O in [3,7)    -> rec2 {spell:B1, type:O, ot:4}
#
# Rule 2 — first_h in [9, 13):
#   if span >= 3       -> rec1 {spell:A2, type:O, ot:4}
#   if span > 4:
#     d1 = last_h - 14
#     if d1 >= 7       -> rec2 {spell:B,  type:R, working:8}
#     elif d1 in [3,7) -> rec2 {spell:B1, type:R, working:4}
#
# Rule 3 — first_h in [13, 21):
#   eff = last_h - 14
#   if eff >= 7        -> rec1 {spell:B,  type:R, working:8}
#   elif eff in [3,7)  -> rec1 {spell:B1, type:R, working:4}
#
# On an off day, the spell + working/ot calculations are unchanged; only the
# attendance_type is forced to "O" on every row.
# ─────────────────────────────────────────────────────────────────────────────

_ETRACK_SPELL_TIMES: dict[str, tuple[str, str]] = {
    "A":  ("06:00:00", "14:00:00"),
    "A1": ("06:00:00", "14:00:00"),
    "A2": ("09:00:00", "13:00:00"),
    "B":  ("14:00:00", "22:00:00"),
    "B1": ("14:00:00", "22:00:00"),
}


def _etrack_records_for_employee(
    first_h: float, last_h: float, span: float,
) -> list[dict]:
    """Apply rules #1/#2/#3 and return a list of {spell, att_type, working, ot}
    dicts (0..2 entries)."""
    out: list[dict] = []

    # Rule 1: first entry [5, 9)
    if 5 <= first_h < 9:
        eff = last_h - 6
        rec1_working = 0.0
        if eff >= 7:
            out.append({"spell": "A",  "att_type": "R", "working": 8.0, "ot": 0.0})
            rec1_working = 8.0
        elif eff >= 3:
            out.append({"spell": "A1", "att_type": "R", "working": 4.0, "ot": 0.0})
            rec1_working = 4.0
        if rec1_working > 0:
            O = eff - rec1_working
            if O >= 7:
                out.append({"spell": "B",  "att_type": "O", "working": 0.0, "ot": 8.0})
            elif O >= 3:
                out.append({"spell": "B1", "att_type": "O", "working": 0.0, "ot": 4.0})

    # Rule 2: first entry [9, 13)
    elif 9 <= first_h < 13:
        if span >= 3:
            out.append({"spell": "A2", "att_type": "O", "working": 0.0, "ot": 4.0})
        if span > 4:
            d1 = last_h - 14
            if d1 >= 7:
                out.append({"spell": "B",  "att_type": "R", "working": 8.0, "ot": 0.0})
            elif d1 >= 3:
                out.append({"spell": "B1", "att_type": "R", "working": 4.0, "ot": 0.0})

    # Rule 3: first entry [13, 21)
    elif 13 <= first_h < 21:
        eff = last_h - 14
        if eff >= 7:
            out.append({"spell": "B",  "att_type": "R", "working": 8.0, "ot": 0.0})
        elif eff >= 3:
            out.append({"spell": "B1", "att_type": "R", "working": 4.0, "ot": 0.0})

    return out


def _process_etrack_day(
    db: Session,
    *,
    tran_date: str,
    is_off_day: bool,
) -> int:
    """Insert daily_attendance_process_table rows for the day per the etrack
    rules. Returns total number of rows inserted (across all employees)."""
    rows = db.execute(
        FETCH_SPELL_PUNCHES_SQL,
        {"tran_date": tran_date, "in_from": "00:00:00", "window_end": "23:59:59"},
    ).fetchall()

    by_emp: dict[int, list[tuple]] = {}
    for r in rows:
        m = r._mapping
        by_emp.setdefault(m["eb_id"], []).append(
            (m["punch_time"], m["bio_att_log_id"], m["dept_id"], m["desig_id"], m["log_date"])
        )

    inserted = 0
    for eb_id, punches in by_emp.items():
        if len(punches) < 2:
            continue
        first_time, first_bio, first_dept, first_desig, first_log_date = punches[0]
        last_time, _, _, _, last_log_date = punches[-1]
        first_sec = _to_seconds(first_time)
        last_sec  = _to_seconds(last_time)
        span_secs = max(0, last_sec - first_sec)
        if span_secs < MIN_PAIR_SECONDS:
            continue

        first_h = first_sec / 3600.0
        last_h  = last_sec  / 3600.0
        span    = round(last_h - first_h, 4)

        recs = _etrack_records_for_employee(first_h, last_h, span)
        if not recs:
            continue

        for rec in recs:
            # On an off day, spell + working/ot stay the same, only the
            # attendance_type is forced to "O".
            att_type      = "O" if is_off_day else rec["att_type"]
            working_hours = rec["working"]
            ot_hours      = rec["ot"]
            time_duration = round(working_hours + ot_hours, 2)
            spell_start, spell_end = _ETRACK_SPELL_TIMES.get(
                rec["spell"], ("00:00:00", "00:00:00"),
            )
            db.execute(
                INSERT_SPELL_ROW_SQL,
                {
                    "eb_id": int(eb_id),
                    "bio_id": int(first_bio) if first_bio is not None else None,
                    "dept_id": int(first_dept) if first_dept is not None else None,
                    "desig_id": int(first_desig) if first_desig is not None else None,
                    "tran_date": tran_date,
                    "spell_name": rec["spell"],
                    "attendance_type": att_type,
                    "check_in": first_log_date,
                    "check_out": last_log_date,
                    "time_duration": time_duration,
                    "working_hours": working_hours,
                    "ot_hours": ot_hours,
                    "spell_start": spell_start,
                    "spell_end": spell_end,
                    "spell_hours": SPELL_HOURS,
                },
            )
            inserted += 1

    print(f"[bio_att_etrack_process]   -> rows inserted = {inserted}", flush=True)
    return inserted


def _build_unmatched_xlsx(rows) -> bytes:
    """Build an xlsx workbook listing the raw bio rows for unmatched emp_codes."""
    try:
        import openpyxl
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl not installed: {e}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unmatched Emp Codes"
    headers = [
        "emp_code", "emp_anme", "bio_id", "log_date", "company_name",
        "department", "designation", "employement_type",
        "device_direction", "device_name",
    ]
    ws.append(headers)
    for r in rows:
        m = r._mapping
        ws.append([
            m.get("emp_code"),
            m.get("emp_anme"),
            m.get("bio_id"),
            str(m.get("log_date") or ""),
            m.get("company_name"),
            m.get("department"),
            m.get("designation"),
            m.get("employement_type"),
            m.get("device_direction"),
            m.get("device_name"),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.post("/bio_att_process")
async def bio_att_process(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Process raw bio_attendance_table rows for a date into
    daily_attendance_process_table.

    Request body: ``{"tran_date": "YYYY-MM-DD"}``.

    Behaviour:
      * If any emp_code on that date is unknown to ``hrms_ed_official_details``,
        responds with an XLSX attachment listing those raw rows and aborts —
        no inserts are made.
      * Otherwise wipes existing rows for that date and re-inserts spell A,
        spell B, and OT rows. Returns JSON stats.
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        body = await request.json()
    except Exception:
        body = {}
    tran_date = (body or {}).get("tran_date")
    if not tran_date:
        raise HTTPException(status_code=400, detail="tran_date is required")
    try:
        # Validate format.
        datetime.strptime(tran_date, "%Y-%m-%d")
    except Exception:
        raise HTTPException(status_code=400, detail="tran_date must be YYYY-MM-DD")

    try:
        # Step 1: collect emp_codes that punched on this date but aren't in the
        # employee master. We don't abort — they're just skipped (the INSERT
        # JOINs against hrms_ed_official_details, so unmatched rows fall out
        # naturally). The list is returned as an xlsx attachment.
        try:
            _log_sql("UNMATCHED_EMP_CODES_SQL", UNMATCHED_EMP_CODES_SQL, {"tran_date": tran_date})
            unmatched = db.execute(
                UNMATCHED_EMP_CODES_SQL, {"tran_date": tran_date}
            ).fetchall()
            print(f"[bio_att_process] unmatched rows: {len(unmatched)}", flush=True)
        except Exception as e:
            traceback.print_exc()
            raise HTTPException(
                status_code=500,
                detail=(
                    "Failed reading bio_attendance_table / hrms_ed_official_details. "
                    f"Original error: {e}"
                ),
            )
        unique_codes = {r._mapping.get("emp_code") for r in unmatched}

        # Step 2: weekly-off check. Fail-soft — if tbl_offday_mst doesn't
        # exist or has unexpected schema, treat the day as a working day.
        try:
            _log_sql("IS_OFF_DAY_SQL", IS_OFF_DAY_SQL, {"tran_date": tran_date})
            off_row = db.execute(IS_OFF_DAY_SQL, {"tran_date": tran_date}).fetchone()
            is_off_day = bool(off_row and int(off_row.cnt) > 0)
            print(f"[bio_att_process] is_off_day = {is_off_day}", flush=True)
        except Exception as e:
            print(f"[bio_att_process] IS_OFF_DAY_SQL failed (treating as working day): {e}", flush=True)
            try:
                db.rollback()
            except Exception:
                pass
            is_off_day = False

        # Step 3: wipe existing rows for this date (re-run safe).
        try:
            _log_sql("DELETE_DAY_ROWS_SQL", DELETE_DAY_ROWS_SQL, {"tran_date": tran_date})
            db.execute(DELETE_DAY_ROWS_SQL, {"tran_date": tran_date})
        except Exception as e:
            traceback.print_exc()
            raise HTTPException(
                status_code=500,
                detail=(
                    "daily_attendance_process_table not accessible. "
                    "Create the table (see DDL in repo notes). "
                    f"Original error: {e}"
                ),
            )

        # Step 4: process Spell A and Spell B for matched employees only.
        try:
            a_reg, a_ot = _process_one_spell(
                db, tran_date=tran_date, spell_name="A",
                in_from=SPELL_A_IN_FROM, in_to=SPELL_A_IN_TO,
                spell_start=SPELL_A_START, spell_end=SPELL_A_END,
                is_off_day=is_off_day,
            )
            b_reg, b_ot = _process_one_spell(
                db, tran_date=tran_date, spell_name="B",
                in_from=SPELL_B_IN_FROM, in_to=SPELL_B_IN_TO,
                spell_start=SPELL_B_START, spell_end=SPELL_B_END,
                is_off_day=is_off_day,
            )
            db.commit()
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            raise HTTPException(
                status_code=500,
                detail=f"Spell processing failed: {e}",
            )

        total = a_reg + a_ot + b_reg + b_ot
        message = (
            f"Processed {total} row(s) for {tran_date} "
            f"(A: {a_reg} regular + {a_ot} OT, "
            f"B: {b_reg} regular + {b_ot} OT, "
            f"day_off: {is_off_day}, "
            f"unmatched_emp_codes: {len(unique_codes)})."
        )

        # If any emp_codes were unmatched, stream the xlsx report alongside
        # success — stats travel in headers so the UI can still show them.
        if unmatched:
            xlsx_bytes = _build_unmatched_xlsx(unmatched)
            return Response(
                content=xlsx_bytes,
                status_code=200,
                media_type=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
                headers={
                    "Content-Disposition": (
                        f'attachment; filename="unmatched_emp_codes_{tran_date}.xlsx"'
                    ),
                    "Access-Control-Expose-Headers": (
                        "Content-Disposition, X-Unmatched-Count, X-Unmatched-Rows, "
                        "X-Processed, X-Process-Message"
                    ),
                    "X-Unmatched-Count": str(len(unique_codes)),
                    "X-Unmatched-Rows": str(len(unmatched)),
                    "X-Processed": str(total),
                    "X-Process-Message": message,
                },
            )

        return {
            "message": message,
            "processed": total,
            "tran_date": tran_date,
            "is_off_day": is_off_day,
            "spell_a_regular": a_reg,
            "spell_a_ot": a_ot,
            "spell_b_regular": b_reg,
            "spell_b_ot": b_ot,
            "unmatched_emp_codes": 0,
        }
    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        # Log the full traceback so we can see it in uvicorn console.
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


# ---------------------------------------------------------------------------
# Final Process: copy daily_attendance_process_table -> daily_attendance.
# ---------------------------------------------------------------------------
# Spell rule (per user spec):
#   * Default: spell A if 06:00 <= ref_time < 14:00, spell B if 14:00 <= ref_time < 22:00.
#   * ref_time = check_out when ot_hours > 0, else check_in.
# Only rows with processed = 1 are copied; copied rows are marked processed = 2.

FINAL_FETCH_SQL = text(
    """
    SELECT daily_att_proc_id, eb_id, bio_id, dept_id, desig_id,
           attendance_date, spell_name, attendance_type, attendance_source,
           check_in, check_out, Time_duration, Working_hours, Ot_hours,
           spell_start_time, spell_end_time, spell_hours, device_name
      FROM daily_attendance_process_table
     WHERE attendance_date = :tran_date
       AND processed = 1
     ORDER BY eb_id, daily_att_proc_id
    """
)

FINAL_INSERT_SQL = text(
    """
    INSERT INTO daily_attendance
      (attendance_date, attendance_source, attendance_type, attendance_mark,
       eb_id, bio_id, branch_id, status_id,
       worked_department_id, worked_designation_id,
       entry_time, exit_time,
       working_hours, idle_hours, spell, spell_hours,
       is_active, update_date_time)
    VALUES
      (:attendance_date, :attendance_source, :attendance_type, 'P',
       :eb_id, :bio_id, :branch_id, 1,
       :worked_department_id, :worked_designation_id,
       :entry_time, :exit_time,
       :working_hours, 0, :spell, :spell_hours,
       1, NOW())
    """
)

FINAL_MARK_PROCESSED_SQL = text(
    """
    UPDATE daily_attendance_process_table
       SET processed = 2
     WHERE attendance_date = :tran_date
       AND processed = 1
    """
)

FINAL_DELETE_EXISTING_SQL = text(
    """
    DELETE FROM daily_attendance
     WHERE bio_id IN :bio_ids
    """
)

# Fetch last daily_ebmc_attendance record for an eb_id, joined to
# daily_attendance so we can compare dept/desig.
FINAL_LAST_EBMC_SQL = text(
    """
    SELECT dea.mc_id,
           da.worked_department_id AS dept_id,
           da.worked_designation_id AS desig_id
    FROM daily_ebmc_attendance dea
    JOIN daily_attendance da ON da.daily_atten_id = dea.daily_atten_id
    WHERE dea.eb_id = :eb_id
      AND COALESCE(dea.is_active, 1) = 1
    ORDER BY dea.dtl_rec_id DESC
    LIMIT 1
    """
)

# Insert one mc entry into daily_ebmc_attendance.
FINAL_INSERT_EBMC_SQL = text(
    """
    INSERT INTO daily_ebmc_attendance
        (daily_atten_id, eb_id, mc_id, is_active)
    VALUES
        (:daily_atten_id, :eb_id, :mc_id, 1)
    """
)

# ── Final-process: update-in-place reconciliation ───────────────────────────
# The final process now writes spell + hours straight from
# daily_attendance_process_table (spell_name / Working_hours / Ot_hours /
# spell_hours) instead of recomputing them. Existing daily_attendance rows for
# the date are UPDATED in place (preserving daily_atten_id and any linked
# daily_ebmc_attendance rows); surplus desired rows are inserted and stale rows
# removed. Rows are paired per employee in deterministic order
# (process rows ORDER BY eb_id, daily_att_proc_id ; existing rows by daily_atten_id).

# Existing BIO daily_attendance rows for a date, for the given employees.
FINAL_SELECT_EXISTING_SQL = text(
    """
    SELECT daily_atten_id, eb_id
      FROM daily_attendance
     WHERE attendance_date = :tran_date
       AND eb_id IN :eb_ids
       AND attendance_source = 'BIO'
     ORDER BY eb_id, daily_atten_id
    """
)

# Update an existing daily_attendance row in place (keeps daily_atten_id).
FINAL_UPDATE_EXISTING_SQL = text(
    """
    UPDATE daily_attendance
       SET attendance_date       = :attendance_date,
           attendance_source     = :attendance_source,
           attendance_type       = :attendance_type,
           attendance_mark       = 'P',
           eb_id                 = :eb_id,
           bio_id                = :bio_id,
           branch_id             = :branch_id,
           status_id             = 1,
           worked_department_id  = :worked_department_id,
           worked_designation_id = :worked_designation_id,
           entry_time            = :entry_time,
           exit_time             = :exit_time,
           working_hours         = :working_hours,
           idle_hours            = 0,
           spell                 = :spell,
           spell_hours           = :spell_hours,
           is_active             = 1,
           update_date_time      = NOW()
     WHERE daily_atten_id = :daily_atten_id
    """
)

# Remove a stale daily_attendance row (and its mc links) that no longer has a
# corresponding process-table row.
FINAL_DELETE_STALE_SQL = text(
    "DELETE FROM daily_attendance WHERE daily_atten_id = :daily_atten_id"
)
FINAL_DELETE_EBMC_BY_ATT_SQL = text(
    "DELETE FROM daily_ebmc_attendance WHERE daily_atten_id = :daily_atten_id"
)


def _build_desired_daily_rows(
    rows,
    *,
    branch_id: int,
    off_eb_ids,
    working_att_type: str,
) -> tuple[dict[int, list[dict]], list[dict], int]:
    """Translate processed daily_attendance_process_table rows into the desired
    daily_attendance rows, taking spell + hours straight from the process table.

    Each process row yields up to two desired rows: a working row
    (attendance_type=`working_att_type`, hours=Working_hours) when Working_hours>0,
    and an OT row (attendance_type='O', hours=Ot_hours) when Ot_hours>0. The
    `spell` comes from the process row's `spell_name` verbatim and `spell_hours`
    from `spell_hours` — no recompute / re-bucketing.

    Returns (desired_by_eb, desired_no_eb, skipped) where desired_by_eb maps
    eb_id -> ordered list of desired-row dicts. `_dept_id`/`_desig_id` private
    keys carry the mc-link source and are stripped before SQL execution.
    """
    desired_by_eb: dict[int, list[dict]] = {}
    desired_no_eb: list[dict] = []
    skipped = 0

    for r in rows:
        m = r._mapping

        spell = m.get("spell_name")
        if spell is None or str(spell).strip() == "":
            # No spell on the process row — cannot place it; skip.
            skipped += 1
            continue

        try:
            wh = float(m.get("Working_hours") or 0)
        except Exception:
            wh = 0.0
        try:
            ot = float(m.get("Ot_hours") or 0)
        except Exception:
            ot = 0.0

        entries: list[tuple[str, float]] = []
        if wh > 0:
            entries.append((working_att_type, wh))
        if ot > 0:
            entries.append(("O", ot))
        if not entries:
            skipped += 1
            continue

        eb_id_val = m.get("eb_id")
        emp_off_day = eb_id_val is not None and int(eb_id_val) in off_eb_ids

        for att_type, hours in entries:
            desired = {
                "attendance_date":       m.get("attendance_date"),
                "attendance_source":     m.get("attendance_source") or "BIO",
                "attendance_type":       "O" if emp_off_day else att_type,
                "eb_id":                 eb_id_val,
                "bio_id":                m.get("bio_id"),
                "branch_id":             branch_id,
                "worked_department_id":  m.get("dept_id"),
                "worked_designation_id": m.get("desig_id"),
                "entry_time":            m.get("check_in"),
                "exit_time":             m.get("check_out"),
                "working_hours":         hours,
                "spell":                 spell,
                "spell_hours":           m.get("spell_hours"),
                # private (mc-link source) — stripped before SQL execution.
                "_dept_id":              m.get("dept_id"),
                "_desig_id":             m.get("desig_id"),
            }
            if eb_id_val is None:
                desired_no_eb.append(desired)
            else:
                desired_by_eb.setdefault(int(eb_id_val), []).append(desired)

    return desired_by_eb, desired_no_eb, skipped


def _maybe_insert_ebmc(db: Session, daily_atten_id, eb_id, dept_id, desig_id) -> None:
    """Insert a daily_ebmc_attendance row iff the employee's last mc entry
    matches the current dept/desig."""
    if not (daily_atten_id and eb_id and dept_id and desig_id):
        return
    last_ebmc = db.execute(FINAL_LAST_EBMC_SQL, {"eb_id": eb_id}).fetchone()
    if (
        last_ebmc is not None
        and last_ebmc.dept_id is not None
        and last_ebmc.desig_id is not None
        and int(last_ebmc.dept_id) == int(dept_id)
        and int(last_ebmc.desig_id) == int(desig_id)
    ):
        db.execute(FINAL_INSERT_EBMC_SQL, {
            "daily_atten_id": daily_atten_id,
            "eb_id": eb_id,
            "mc_id": last_ebmc.mc_id,
        })


def _insert_daily_att_row(db: Session, desired: dict) -> int | None:
    """Insert one daily_attendance row from a desired-row dict and add its mc
    link if applicable. Returns the new daily_atten_id."""
    params = {k: v for k, v in desired.items() if not k.startswith("_")}
    ins_res = db.execute(FINAL_INSERT_SQL, params)
    new_id = ins_res.lastrowid
    _maybe_insert_ebmc(db, new_id, desired.get("eb_id"),
                       desired.get("_dept_id"), desired.get("_desig_id"))
    return new_id


def _update_daily_att_row(db: Session, daily_atten_id: int, desired: dict) -> None:
    """Update an existing daily_attendance row in place from a desired-row dict.
    The mc link is left intact (daily_atten_id is preserved)."""
    params = {k: v for k, v in desired.items() if not k.startswith("_")}
    params["daily_atten_id"] = daily_atten_id
    db.execute(FINAL_UPDATE_EXISTING_SQL, params)


def finalize_daily_attendance(
    db: Session,
    rows,
    *,
    tran_date: str,
    branch_id: int,
    off_eb_ids=frozenset(),
    working_att_type: str = "R",
) -> dict:
    """Reconcile daily_attendance for `tran_date` from processed process-table
    `rows`, taking spell + hours from the process table.

    Existing BIO rows for the date are UPDATED in place (preserving
    daily_atten_id and linked daily_ebmc_attendance); extra desired rows are
    inserted and stale rows deleted. Does NOT commit — the caller commits.

    Returns dict(inserted, updated, deleted, skipped).
    """
    desired_by_eb, desired_no_eb, skipped = _build_desired_daily_rows(
        rows, branch_id=branch_id, off_eb_ids=off_eb_ids,
        working_att_type=working_att_type,
    )

    existing_by_eb: dict[int, list[int]] = {}
    eb_ids = list(desired_by_eb.keys())
    if eb_ids:
        ex_rows = db.execute(
            FINAL_SELECT_EXISTING_SQL,
            {"tran_date": tran_date, "eb_ids": tuple(eb_ids)},
        ).fetchall()
        for er in ex_rows:
            em = er._mapping
            existing_by_eb.setdefault(int(em["eb_id"]), []).append(
                int(em["daily_atten_id"])
            )

    inserted = updated = deleted = 0

    # Rows with no eb_id can't be matched — always insert fresh.
    for desired in desired_no_eb:
        _insert_daily_att_row(db, desired)
        inserted += 1

    for eb_id, desired_rows in desired_by_eb.items():
        existing_ids = existing_by_eb.get(eb_id, [])
        n = max(len(desired_rows), len(existing_ids))
        for i in range(n):
            if i < len(desired_rows) and i < len(existing_ids):
                _update_daily_att_row(db, existing_ids[i], desired_rows[i])
                updated += 1
            elif i < len(desired_rows):
                _insert_daily_att_row(db, desired_rows[i])
                inserted += 1
            else:
                # Stale row — no longer produced by the process table.
                db.execute(FINAL_DELETE_EBMC_BY_ATT_SQL,
                           {"daily_atten_id": existing_ids[i]})
                db.execute(FINAL_DELETE_STALE_SQL,
                           {"daily_atten_id": existing_ids[i]})
                deleted += 1

    return {
        "inserted": inserted,
        "updated": updated,
        "deleted": deleted,
        "skipped": skipped,
    }


def _resolve_spell_by_time(check_in, check_out, ot_hours) -> str | None:
    """Return 'A' / 'B' based on time-of-day rule.

    ref = check_out when ot_hours > 0 else check_in.
    A: 06:00 <= ref < 14:00 ; otherwise B.
    """
    try:
        ot_val = float(ot_hours or 0)
    except Exception:
        ot_val = 0.0
    ref = check_out if ot_val > 0 else check_in
    if ref is None:
        return None
    # Normalise to time-of-day seconds.
    if isinstance(ref, datetime):
        secs = ref.hour * 3600 + ref.minute * 60 + ref.second
    elif isinstance(ref, dt_time):
        secs = ref.hour * 3600 + ref.minute * 60 + ref.second
    elif isinstance(ref, timedelta):
        total = int(ref.total_seconds()) % 86400
        secs = total
    else:
        # Last-ditch: try to parse "HH:MM[:SS]".
        try:
            parts = str(ref).split(":")
            secs = int(parts[0]) * 3600 + int(parts[1]) * 60 + (int(parts[2]) if len(parts) > 2 else 0)
        except Exception:
            return None
    if 6 * 3600 <= secs < 14 * 3600:
        return "A"
    return "B"


@router.post("/bio_att_final_process")
async def bio_att_final_process(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Move processed=1 rows from daily_attendance_process_table into
    daily_attendance for a given date, recomputing spell by time-of-day.
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        body = await request.json()
    except Exception:
        body = {}
    tran_date = (body or {}).get("tran_date")
    if not tran_date:
        raise HTTPException(status_code=400, detail="tran_date is required")
    try:
        datetime.strptime(tran_date, "%Y-%m-%d")
    except Exception:
        raise HTTPException(status_code=400, detail="tran_date must be YYYY-MM-DD")

    branch_id_raw = (body or {}).get("branch_id")
    if branch_id_raw in (None, ""):
        raise HTTPException(status_code=400, detail="branch_id is required")
    try:
        branch_id = int(branch_id_raw)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="branch_id must be an integer")

    try:
        # Per-employee off-day set — eb_ids whose weekly off matches
        # DAYOFWEEK(tran_date). Any daily_attendance row written for one of
        # these employees on this date gets attendance_type forced to 'O'.
        off_eb_rows = db.execute(
            text(
                "SELECT eb_id FROM tbl_offday_mst "
                "WHERE off_day = DAYOFWEEK(:tran_date) AND eb_id IS NOT NULL"
            ),
            {"tran_date": tran_date},
        ).fetchall()
        off_eb_ids: set[int] = {
            int(r._mapping["eb_id"]) for r in off_eb_rows
        }
        print(
            f"[bio_att_final_process] off-day eb_ids count = {len(off_eb_ids)} "
            f"for tran_date={tran_date}",
            flush=True,
        )

        _log_sql("FINAL_FETCH_SQL", FINAL_FETCH_SQL, {"tran_date": tran_date})
        rows = db.execute(FINAL_FETCH_SQL, {"tran_date": tran_date}).fetchall()
        print(f"[bio_att_final_process] fetched {len(rows)} row(s)", flush=True)

        # Write spell + hours straight from the process table; UPDATE existing
        # daily_attendance rows in place (preserving daily_atten_id / mc links),
        # insert surplus rows, remove stale ones.
        stats = finalize_daily_attendance(
            db, rows,
            tran_date=tran_date,
            branch_id=branch_id,
            off_eb_ids=off_eb_ids,
            working_att_type="R",
        )
        print(
            f"[bio_att_final_process] {tran_date} -> inserted={stats['inserted']} "
            f"updated={stats['updated']} deleted={stats['deleted']} "
            f"skipped={stats['skipped']}",
            flush=True,
        )

        _log_sql("FINAL_MARK_PROCESSED_SQL", FINAL_MARK_PROCESSED_SQL, {"tran_date": tran_date})
        db.execute(FINAL_MARK_PROCESSED_SQL, {"tran_date": tran_date})
        db.commit()

        return {
            "message": (
                f"Final processed {tran_date}: inserted {stats['inserted']}, "
                f"updated {stats['updated']}, deleted {stats['deleted']} "
                f"(skipped {stats['skipped']})."
            ),
            "tran_date": tran_date,
            "inserted": stats["inserted"],
            "updated": stats["updated"],
            "deleted": stats["deleted"],
            "skipped": stats["skipped"],
        }
    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


# ---------------------------------------------------------------------------
# Etrack SQL Server -> bio_attendance_table transfer
# ---------------------------------------------------------------------------

ETRACK_INSERT_SQL = text(
    """
    INSERT INTO bio_attendance_table
        (bio_att_log_id, emp_code, emp_anme, bio_id, log_date,
         device_direction, device_id)
    VALUES (:bio_att_log_id, :emp_code, :emp_anme, :bio_id, :log_date,
            :device_direction, :device_id)
    """
)


@router.post("/bio_att_etrack")
async def bio_att_etrack(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Transfer of punches from the Etrack SQL Server to bio_attendance_table.

    Two modes:

    1. Single-day mode — when `tran_date` is supplied:
           Fetches every row whose CONVERT(DATE, dl.LogDate) = tran_date from
           that day's monthly DeviceLogs_<m>_<y> table. No lower bound from
           prior records is applied — the entire selected day is pulled.

    2. Auto-incremental mode — when `tran_date` is omitted:
           The transfer range starts from MAX(log_date) in bio_attendance_table
           and spans every monthly DeviceLogs_<m>_<y> table from that month up
           to today's month, capped at dl.LogDate < (today + 1).
               - Start month  : dl.LogDate > last_log_date
               - Later months : all rows (still capped at today + 1)
           If bio_attendance_table is empty, only today's month is pulled.

    Body / query param (optional):
        company_id : Etrack CompanyId to filter on (default = 2)
        tran_date  : YYYY-MM-DD — switches to single-day mode

    De-dup key on the MySQL side: bio_att_log_id (= SQL Server DeviceLogId).
    """
    try:
        body: dict = {}
        try:
            body = await request.json()
        except Exception:
            body = {}
        qp = request.query_params

        company_id_raw = body.get("company_id") or qp.get("company_id") or "2"
        try:
            company_id = int(company_id_raw)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Invalid company_id")

        # Optional override for the upper bound of the transfer range. When
        # omitted or blank, falls back to today.
        tran_date_raw = body.get("tran_date") or qp.get("tran_date")
        tran_date_override: date | None = None
        if tran_date_raw:
            try:
                tran_date_override = datetime.strptime(
                    str(tran_date_raw), "%Y-%m-%d",
                ).date()
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid tran_date {tran_date_raw!r}, expected YYYY-MM-DD",
                )

        # Lazy import so the rest of the module still works without pyodbc.
        try:
            from src.hrms.etrack_conn import (
                device_logs_table_name,
                get_etrack_connection,
            )
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Etrack connector not available: {e}",
            )

        # ------------------------------------------------------------------
        # Determine the auto-incremental start point.
        # ------------------------------------------------------------------
        today = tran_date_override or date.today()

        last_log_row = db.execute(
            text(
                "SELECT MAX(log_date) AS max_log_date "
                "FROM bio_attendance_table"
            )
        ).mappings().first()
        last_log_date_dt = last_log_row["max_log_date"] if last_log_row else None

        if last_log_date_dt is None:
            # No prior data: start from selected day's month, take everything
            # in it (still capped at end_exclusive below).
            start_month_date = today.replace(day=1)
            last_log_dt: datetime | None = None
            from_log_date_iso: str | None = None
        else:
            if isinstance(last_log_date_dt, datetime):
                last_log_dt = last_log_date_dt
                last_log_d = last_log_date_dt.date()
            else:
                last_log_dt = datetime.combine(last_log_date_dt, dt_time.min)
                last_log_d = last_log_date_dt
            start_month_date = last_log_d.replace(day=1)
            from_log_date_iso = (
                last_log_date_dt.isoformat()
                if hasattr(last_log_date_dt, "isoformat")
                else str(last_log_date_dt)
            )

        # When a specific tran_date is supplied, fetch only that single day
        # from its monthly table, filtered by dl.DeviceLogId > last_id_for_day
        # so we incrementally pick up new punches for that date.
        # Otherwise, build the list of months from start_month_date through
        # today and let the per-month filters (last_log_dt + end_exclusive)
        # define the range.
        single_day_mode = tran_date_override is not None
        last_log_id_for_day = 0
        if single_day_mode:
            id_row = db.execute(
                text(
                    "SELECT MAX(bio_att_log_id) AS max_id "
                    "FROM bio_attendance_table "
                    "WHERE DATE(log_date) = :d"
                ),
                {"d": today},
            ).mappings().first()
            last_log_id_for_day = (
                int(id_row["max_id"])
                if id_row and id_row["max_id"] is not None
                else 0
            )

        months: list[date] = []
        if single_day_mode:
            months.append(today.replace(day=1))
        else:
            cursor_month = start_month_date
            end_month = today.replace(day=1)
            while cursor_month <= end_month:
                months.append(cursor_month)
                if cursor_month.month == 12:
                    cursor_month = date(cursor_month.year + 1, 1, 1)
                else:
                    cursor_month = date(
                        cursor_month.year, cursor_month.month + 1, 1
                    )

        # ------------------------------------------------------------------
        # Fetch from each monthly DeviceLogs table.
        # ------------------------------------------------------------------
        try:
            sconn = get_etrack_connection()
        except Exception as e:
            raise HTTPException(
                status_code=502,
                detail=f"Cannot connect to Etrack SQL Server: {e}",
            )

        total_fetched = 0
        total_inserted = 0
        per_table: list[dict] = []

        # Upper-bound LogDate (exclusive) — caps the transfer at the end of
        # the selected day (tran_date_override) or today.
        end_exclusive = today + timedelta(days=1)

        # ------------------------------------------------------------------
        # CSV dump (DISABLED). One file per request under exports/etrack/.
        # ------------------------------------------------------------------
        # co_id_for_path = (qp.get("co_id") or "unknown").replace("/", "_")
        # csv_dir = os.path.join("exports", "etrack")
        # os.makedirs(csv_dir, exist_ok=True)
        # csv_filename = (
        #     f"{co_id_for_path}_{today.isoformat()}_"
        #     f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        # )
        # csv_path = os.path.abspath(os.path.join(csv_dir, csv_filename))
        # csv_file = open(csv_path, "w", newline="", encoding="utf-8")
        # csv_writer = csv.writer(csv_file)
        # csv_writer.writerow([
        #     "DeviceLogId", "DeviceId", "UserId", "LogDate", "Direction",
        #     "EmployeeId", "EmployeeCode", "EmployeeName", "CompanyId",
        #     "source_table",
        # ])

        try:
            cur = sconn.cursor()
            for idx, m_date in enumerate(months):
                table_name = device_logs_table_name(m_date)
                is_start_month = (idx == 0)

                if single_day_mode:
                    sql = (
                        f"SELECT dl.DeviceLogId, dl.DeviceId, dl.UserId, dl.LogDate, "
                        f"       dl.Direction, em.EmployeeId, em.EmployeeCode, "
                        f"       em.EmployeeName, em.CompanyId "
                        f"FROM dbo.{table_name} dl "
                        f"LEFT JOIN dbo.Employees em "
                        f"  ON em.EmployeeCodeInDevice = dl.UserId "
                        f"WHERE em.CompanyId = ? "
                        f"  AND dl.DeviceLogId > ? "
                        f"  AND CONVERT(DATE, dl.LogDate) = ?"
                    )
                    params_args: tuple = (company_id, last_log_id_for_day, today)
                elif is_start_month and last_log_dt is not None:
                    sql = (
                        f"SELECT dl.DeviceLogId, dl.DeviceId, dl.UserId, dl.LogDate, "
                        f"       dl.Direction, em.EmployeeId, em.EmployeeCode, "
                        f"       em.EmployeeName, em.CompanyId "
                        f"FROM dbo.{table_name} dl "
                        f"LEFT JOIN dbo.Employees em "
                        f"  ON em.EmployeeCodeInDevice = dl.UserId "
                        f"WHERE dl.LogDate > ? "
                        f"  AND em.CompanyId = ? "
                        f"  AND dl.LogDate < ?"
                    )
                    params_args = (last_log_dt, company_id, end_exclusive)
                else:
                    sql = (
                        f"SELECT dl.DeviceLogId, dl.DeviceId, dl.UserId, dl.LogDate, "
                        f"       dl.Direction, em.EmployeeId, em.EmployeeCode, "
                        f"       em.EmployeeName, em.CompanyId "
                        f"FROM dbo.{table_name} dl "
                        f"LEFT JOIN dbo.Employees em "
                        f"  ON em.EmployeeCodeInDevice = dl.UserId "
                        f"WHERE em.CompanyId = ? "
                        f"  AND dl.LogDate < ?"
                    )
                    params_args = (company_id, end_exclusive)
                try:
                    cur.execute(sql, *params_args)
                    src_rows = cur.fetchall()
                except Exception as e:
                    raise HTTPException(
                        status_code=500,
                        detail=f"Etrack query failed on {table_name}: {e}",
                    )
                print(sql)
                print(f"[bio_att_etrack] Executed on {table_name} with params {params_args}", flush=True)
                fetched = len(src_rows)
                print(f"[bio_att_etrack]  == {table_name}: fetched {fetched} row(s)", flush=True)
                inserted = 0
                BATCH_SIZE = 500
                batch: list[dict] = []

                def _flush_batch() -> int:
                    if not batch:
                        return 0
                    try:
                        res = db.execute(ETRACK_INSERT_SQL, batch)
                        flushed = int(res.rowcount or 0)
                        if flushed < 0:
                            flushed = len(batch)
                    except Exception as e:
                        print(
                            f"[bio_att_etrack] batch insert failed ({len(batch)} rows) on "
                            f"{table_name}: {e}. Falling back to row-by-row.",
                            flush=True,
                        )
                        flushed = 0
                        for p in batch:
                            try:
                                r2 = db.execute(ETRACK_INSERT_SQL, p)
                                flushed += int(r2.rowcount or 0)
                            except Exception as e2:
                                print(
                                    f"[bio_att_etrack] insert failed for "
                                    f"DeviceLogId={p['bio_att_log_id']}: {e2}",
                                    flush=True,
                                )
                    batch.clear()
                    return flushed

                for r in src_rows:
                    # CSV: dump every fetched row as-is (raw values). DISABLED.
                    # csv_writer.writerow([
                    #     r.DeviceLogId, r.DeviceId, r.UserId, r.LogDate,
                    #     r.Direction, r.EmployeeId, r.EmployeeCode,
                    #     r.EmployeeName, r.CompanyId, table_name,
                    # ])

                    params = {
                        "bio_att_log_id": r.DeviceLogId,
                        "emp_code": r.EmployeeCode,
                        "emp_anme": r.EmployeeName,
                        "bio_id": r.UserId,
                        "log_date": r.LogDate,
                        "device_direction": r.Direction,
                        "device_id": r.DeviceId,
                    }
                    if params["bio_att_log_id"] is None:
                        continue
                    batch.append(params)
                    if len(batch) >= BATCH_SIZE:
                        inserted += _flush_batch()

                inserted += _flush_batch()
                print(f"[bio_att_etrack] {table_name}: inserted {inserted} row(s)", flush=True)
                total_fetched += fetched
                total_inserted += inserted
                if single_day_mode:
                    per_from = today.isoformat()
                elif is_start_month:
                    per_from = from_log_date_iso
                else:
                    per_from = None
                per_table.append({
                    "table": table_name,
                    "from_log_date": per_from,
                    "fetched": fetched,
                    "inserted": inserted,
                })
        finally:
            try:
                sconn.close()
            except Exception:
                pass
            # try:
            #     csv_file.close()
            # except Exception:
            #     pass

        db.commit()
        duplicates = max(total_fetched - total_inserted, 0)

        # Step 1: back-fill eb_id (match_type='E') and device_id (match_type='B')
        # from tbl_master_bio_link_mst — same approach as the Excel upload flow.
        link_counts: dict = {}
        try:
            link_counts = _backfill_links(db)
            db.commit()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
            link_counts = {"errors": ["backfill failed"]}

        # Step 2: fill dept_id / desig_id from daily_attendance last record,
        # fallback to hrms_ed_official_details, for any rows with eb_id set
        # but dept_id / desig_id still NULL after step 1.
        dept_desig: dict = {}
        try:
            dept_desig = _resolve_dept_desig(db)
            db.commit()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
            dept_desig = {"errors": ["dept_desig resolve failed"]}

        # First entry of `per_table` corresponds to the start month — keep
        # `source_table` for backwards-compatibility with existing UI.
        source_table = per_table[0]["table"] if per_table else ""

        return {
            "status": "ok",
            "from_log_date": (
                today.isoformat() if single_day_mode else from_log_date_iso
            ),
            "tran_date": today.isoformat(),
            "company_id": company_id,
            "source_table": source_table,
            "tables": per_table,
            "fetched": total_fetched,
            "inserted": total_inserted,
            "duplicates": duplicates,
            # "csv_path": csv_path,  # CSV dump disabled
            "linked": {
                "eb_id": link_counts.get("E", 0),
                "device_id": link_counts.get("B", 0),
                "errors": link_counts.get("errors", []),
            },
            "dept_desig": {
                "candidates": dept_desig.get("candidates", 0),
                "updated": dept_desig.get("updated", 0),
                "from_daily_attendance": dept_desig.get("from_daily_attendance", 0),
                "fallback_official": dept_desig.get("fallback_official", 0),
                "no_source": dept_desig.get("no_source", 0),
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


# ---------------------------------------------------------------------------
# Etrack Process: resolve eb_id / dept_id / desig_id for bio_attendance rows
# where eb_id IS NULL, using:
#   1. tbl_master_bio_link_mst (match_type='E') to resolve emp_code -> eb_id
#   2. last record in daily_attendance for that eb_id
#        -> worked_department_id (dept_id), worked_designation_id (desig_id)
#   3. fallback: hrms_ed_official_details
#        -> sub_dept_id (dept_id), designation_id (desig_id)
# ---------------------------------------------------------------------------

_ETRACK_PROC_UNRESOLVED_SQL = text(
    """
    SELECT DISTINCT b.emp_code, CAST(m.master_id AS SIGNED) AS eb_id
    FROM bio_attendance_table b
    JOIN tbl_master_bio_link_mst m
        ON m.match_type = 'E' AND m.bio_data = b.emp_code
    WHERE b.eb_id IS NULL
      AND b.emp_code IS NOT NULL AND b.emp_code <> ''
    """
)

_ETRACK_PROC_LAST_DAILY_ATT_SQL = text(
    """
    SELECT da.eb_id,
           da.worked_department_id,
           da.worked_designation_id
    FROM daily_attendance da
    INNER JOIN (
        SELECT eb_id, MAX(attendance_date) AS last_date
        FROM daily_attendance
        WHERE eb_id IN :eb_ids
          AND worked_department_id IS NOT NULL
        GROUP BY eb_id
    ) lx ON lx.eb_id = da.eb_id AND lx.last_date = da.attendance_date
    WHERE da.eb_id IN :eb_ids
      AND da.worked_department_id IS NOT NULL
    """
)

_ETRACK_PROC_OFFICIAL_SQL = text(
    """
    SELECT eb_id, sub_dept_id AS dept_id, designation_id AS desig_id
    FROM hrms_ed_official_details
    WHERE eb_id IN :eb_ids
      AND active = 1
    """
)

_ETRACK_PROC_UPDATE_SQL = text(
    """
    UPDATE bio_attendance_table
       SET eb_id    = :eb_id,
           dept_id  = :dept_id,
           desig_id = :desig_id
     WHERE emp_code = :emp_code
       AND eb_id IS NULL
    """
)

# Pre-step (Etrack Process): back-fill emp_code on bio_attendance_table from the
# link master by matching the device bio id. tbl_master_bio_link_mst holds
# bio_dev_id (the device-side bio id) and master_data (the master emp_code) for
# match_type='E'. Overwrites emp_code for every matching row so the subsequent
# emp_code -> eb_id resolution works off the corrected codes.
_ETRACK_PROC_EMPCODE_UPDATE_SQL = text(
    """
    UPDATE bio_attendance_table b
    JOIN tbl_master_bio_link_mst m
        ON m.match_type = 'E'
       AND m.bio_dev_id = b.bio_id
       SET b.emp_code = m.master_data
    """
)

# Pre-step (Etrack Process): fill a blank bio_att_log_id with a generated local
# sequence. Base = 500000 when the table max is below 500000, otherwise
# max + 1; each blank row (NULL or 0) then takes the next sequential id in
# bio_att_id order. @ba_seq is initialised on the same connection just before.
_ETRACK_PROC_LOGID_INIT_SQL = text("SET @ba_seq := :base - 1")

_ETRACK_PROC_LOGID_FILL_SQL = text(
    """
    UPDATE bio_attendance_table
       SET bio_att_log_id = (@ba_seq := @ba_seq + 1)
     WHERE bio_att_log_id IS NULL OR bio_att_log_id = 0
     ORDER BY bio_att_id
    """
)

# Pre-step (Etrack Process): set device_id from the punch direction.
# 'in' -> 22, anything else (out/blank) -> 14. Compared case-insensitively
# since device_direction is stored lowercase ('in'/'out') but may arrive
# capitalised from the device feed.
_ETRACK_PROC_DEVICE_ID_SQL = text(
    """
    UPDATE bio_attendance_table
       SET device_id = CASE
           WHEN LOWER(TRIM(device_direction)) = 'in' THEN 22
           ELSE 14
       END
    """
)

# ---------------------------------------------------------------------------
# Resolve dept_id / desig_id for rows that already have eb_id set but are
# still missing dept_id or desig_id.
# Priority:
#   1. last record in daily_attendance (worked_department_id, worked_designation_id)
#   2. fallback: hrms_ed_official_details (sub_dept_id, designation_id)
# ---------------------------------------------------------------------------

_RESOLVE_DEPT_DESIG_TARGET_SQL = text(
    """
    SELECT DISTINCT b.emp_code, CAST(b.eb_id AS SIGNED) AS eb_id
    FROM bio_attendance_table b
    WHERE b.eb_id IS NOT NULL
      AND (b.dept_id IS NULL OR b.desig_id IS NULL)
    """
)

_RESOLVE_DEPT_DESIG_UPDATE_SQL = text(
    """
    UPDATE bio_attendance_table
       SET dept_id  = :dept_id,
           desig_id = :desig_id
     WHERE emp_code = :emp_code
       AND eb_id    = :eb_id
       AND (dept_id IS NULL OR desig_id IS NULL)
    """
)


def _resolve_dept_desig(db: Session) -> dict:
    """Fill dept_id / desig_id on bio_attendance rows that have eb_id set
    but still have dept_id or desig_id NULL.

    Priority:
      1. Last record in daily_attendance per eb_id
           -> worked_department_id (dept_id), worked_designation_id (desig_id)
      2. Fallback: hrms_ed_official_details
           -> sub_dept_id (dept_id), designation_id (desig_id)

    Returns a stats dict with keys:
        candidates, updated, from_daily_attendance, fallback_official, no_source
    """
    result: dict = {
        "candidates": 0,
        "updated": 0,
        "from_daily_attendance": 0,
        "fallback_official": 0,
        "no_source": 0,
    }

    target_rows = db.execute(_RESOLVE_DEPT_DESIG_TARGET_SQL).fetchall()
    if not target_rows:
        return result

    emp_eb_map: dict[str, int] = {
        str(r.emp_code): int(r.eb_id)
        for r in target_rows
        if r.eb_id is not None
    }
    if not emp_eb_map:
        return result

    result["candidates"] = len(emp_eb_map)
    unique_eb_ids = list(set(emp_eb_map.values()))

    # Step 1: last daily_attendance record per eb_id
    da_rows = db.execute(
        _ETRACK_PROC_LAST_DAILY_ATT_SQL,
        {"eb_ids": tuple(unique_eb_ids)},
    ).fetchall()
    da_map: dict[int, tuple] = {
        int(r.eb_id): (r.worked_department_id, r.worked_designation_id)
        for r in da_rows
        if r.worked_department_id is not None
    }
    result["from_daily_attendance"] = len(da_map)

    # Step 2: fallback from hrms_ed_official_details
    missing_eb_ids = [eid for eid in unique_eb_ids if eid not in da_map]
    official_map: dict[int, tuple] = {}
    if missing_eb_ids:
        off_rows = db.execute(
            _ETRACK_PROC_OFFICIAL_SQL,
            {"eb_ids": tuple(missing_eb_ids)},
        ).fetchall()
        official_map = {
            int(r.eb_id): (r.dept_id, r.desig_id)
            for r in off_rows
        }

    updated = 0
    fallback_official = 0
    no_source = 0

    for emp_code, eb_id in emp_eb_map.items():
        if eb_id in da_map:
            dept_id, desig_id = da_map[eb_id]
        elif eb_id in official_map:
            dept_id, desig_id = official_map[eb_id]
            fallback_official += 1
        else:
            no_source += 1
            continue

        res = db.execute(
            _RESOLVE_DEPT_DESIG_UPDATE_SQL,
            {
                "dept_id": dept_id,
                "desig_id": desig_id,
                "emp_code": emp_code,
                "eb_id": eb_id,
            },
        )
        updated += int(res.rowcount or 0)

    result["updated"] = updated
    result["fallback_official"] = fallback_official
    result["no_source"] = no_source
    return result


def _resolve_eb_id_from_link_master(db: Session) -> dict:
    """Resolve eb_id / dept_id / desig_id for bio_attendance_table rows where
    eb_id IS NULL — the same process the etrack flow uses:

      1. emp_code -> eb_id via tbl_master_bio_link_mst (match_type='E').
      2. dept_id / desig_id from the last daily_attendance record per eb_id.
      3. Fallback dept_id / desig_id from hrms_ed_official_details (active=1).
      4. UPDATE bio_attendance_table SET eb_id, dept_id, desig_id
         WHERE emp_code = ? AND eb_id IS NULL.

    Caller is responsible for committing. Returns a stats dict:
        resolved, updated, from_daily_attendance, fallback_official, no_source.
    """
    result: dict = {
        "resolved": 0,
        "updated": 0,
        "from_daily_attendance": 0,
        "fallback_official": 0,
        "no_source": 0,
    }

    unresolved_rows = db.execute(_ETRACK_PROC_UNRESOLVED_SQL).fetchall()
    if not unresolved_rows:
        return result

    emp_eb_map: dict[str, int] = {
        str(r.emp_code): int(r.eb_id)
        for r in unresolved_rows
        if r.eb_id is not None
    }
    if not emp_eb_map:
        return result

    unique_eb_ids = list(set(emp_eb_map.values()))

    # dept_id / desig_id from the last daily_attendance record per eb_id.
    da_rows = db.execute(
        _ETRACK_PROC_LAST_DAILY_ATT_SQL,
        {"eb_ids": tuple(unique_eb_ids)},
    ).fetchall()
    da_map: dict[int, tuple] = {
        int(r.eb_id): (r.worked_department_id, r.worked_designation_id)
        for r in da_rows
        if r.worked_department_id is not None
    }

    # Fallback dept_id / desig_id from hrms_ed_official_details.
    missing_eb_ids = [eid for eid in unique_eb_ids if eid not in da_map]
    official_map: dict[int, tuple] = {}
    if missing_eb_ids:
        off_rows = db.execute(
            _ETRACK_PROC_OFFICIAL_SQL,
            {"eb_ids": tuple(missing_eb_ids)},
        ).fetchall()
        official_map = {
            int(r.eb_id): (r.dept_id, r.desig_id)
            for r in off_rows
        }

    updated = 0
    fallback_official = 0
    no_source = 0
    for emp_code, eb_id in emp_eb_map.items():
        if eb_id in da_map:
            dept_id, desig_id = da_map[eb_id]
        elif eb_id in official_map:
            dept_id, desig_id = official_map[eb_id]
            fallback_official += 1
        else:
            no_source += 1
            continue

        res = db.execute(
            _ETRACK_PROC_UPDATE_SQL,
            {
                "eb_id": eb_id,
                "dept_id": dept_id,
                "desig_id": desig_id,
                "emp_code": emp_code,
            },
        )
        updated += int(res.rowcount or 0)

    result.update({
        "resolved": len(emp_eb_map),
        "updated": updated,
        "from_daily_attendance": len(da_map),
        "fallback_official": fallback_official,
        "no_source": no_source,
    })
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Bprocess processing. Used by /bio_att_bprocess.
#
# Mirrors eSSL Daily Attendance Detailed Report:
#   Shift   | Sched In | Sched Out | First-IN band
#   --------+----------+-----------+--------------
#   A       | 06:00    | 14:00     | [04:00, 09:00)
#   GS      | 10:00    | 18:00     | [09:00, 13:00)
#   B       | 14:00    | 22:00     | [13:00, 18:00)
#   C       | 22:00    | 06:00 (+1)| [21:00, 24:00)  — last OUT on tran_date+1
#   NS      | -        | -         | no punches → Absent
#
# Per employee/day:
#   - actual_in        = first IN on tran_date
#   - actual_out       = last OUT (on tran_date for A/B/GS; on tran_date+1 for C)
#   - total_dur        = actual_out - actual_in (minutes)
#   - work_dur         = min(8h, total_dur)  -- 8h regular cap
#   - ot               = max(0, total_dur - 8h)
#   - late_by          = max(0, actual_in - sched_in)
#   - early_going_by   = max(0, sched_out - actual_out)
#   - status           = "Present" / "Present (No OutPunch)" / "Absent"
#
# When only an IN exists (No OutPunch):
#   actual_out = scheduled out, work_dur = sched_out - actual_in (capped at 8h),
#   ot = 0, status = "Present (No OutPunch)".
# ─────────────────────────────────────────────────────────────────────────────

_BPROCESS_SPELL_TIMES: dict[str, tuple[str, str]] = {
    "A":  ("06:00:00", "14:00:00"),
    "A1": ("06:00:00", "14:00:00"),
    "A2": ("09:00:00", "13:00:00"),
    "GS": ("10:00:00", "18:00:00"),
    "B":  ("14:00:00", "22:00:00"),
    "B1": ("14:00:00", "22:00:00"),
    "B2": ("17:00:00", "22:00:00"),
    "C":  ("22:00:00", "06:00:00"),
    "C1": ("22:00:00", "06:00:00"),
    "C2": ("00:00:00", "06:00:00"),
    "O":  ("00:00:00", "00:00:00"),
}


def _minutes_bucket(mins: int) -> int:
    """Bucket per spec: >420 -> 8, [180, 420] -> 4, (0, 180) -> 0, <=0 -> -1."""
    if mins <= 0:
        return -1
    if mins > 420:
        return 8
    if mins >= 180:
        return 4
    return 0


def _spell_label_for_b_atten(
    shift: str, kind: str, bucket: int, intime_h: float,
) -> str:
    """Spell label per spellcalculation.txt. `kind` is 'R' or 'O'."""
    if shift == "A":
        if kind == "R":
            if bucket == 8:
                return "A"
            if bucket == 4:
                return "A1" if intime_h < 9 else "A2"
            return "A"
        # OT
        if bucket == 8:
            return "B"
        if bucket == 4:
            return "B1"
        return "O"
    if shift == "B":
        if kind == "R":
            if bucket == 8:
                return "B"
            if bucket == 4:
                return "B1" if intime_h < 17 else "B2"
            return "B"
        # OT
        if bucket == 8:
            return "C"
        if bucket == 4:
            return "C1"
        return "O"
    if shift == "C":
        if kind == "R":
            if bucket == 8:
                return "C"
            if bucket == 4:
                if 0 <= intime_h <= 4:
                    return "C2"
                return "C1"
            return "C"
        # OT — split by intime band.
        if bucket == 8 and intime_h > 21:
            return "A"
        if bucket == 4:
            if intime_h < 19:
                return "B2"
            if intime_h > 21:
                return "A1"
        return "O"
    # GS / unknown shifts — keep simple labels.
    return shift if kind == "R" else "O"


def _bprocess_shift_for(first_h: float) -> str | None:
    """Map first-IN hour to the eSSL shift label, or None for 'NS'/no shift."""
    if 4 <= first_h < 9:
        return "A"
    if 9 <= first_h < 13:
        return "GS"
    if 13 <= first_h < 21:
        return "B"
    if first_h >= 21:
        return "C"
    return None


_REGULAR_WORK_SECONDS = 8 * 3600  # eSSL regular cap (8 hours)


# ─────────────────────────────────────────────────────────────────────────────
# daily_attendance_basic — eSSL-style daily basic report row per employee/day.
# Created lazily (CREATE TABLE IF NOT EXISTS) to avoid a separate migration.
# ─────────────────────────────────────────────────────────────────────────────

_CREATE_DAILY_BASIC_SQL = text(
    """
    CREATE TABLE IF NOT EXISTS daily_attendance_basic (
        id                  INT AUTO_INCREMENT PRIMARY KEY,
        eb_id               INT       NOT NULL,
        emp_code            VARCHAR(32) NULL,
        bio_id              INT       NULL,
        dept_id             INT       NULL,
        desig_id            INT       NULL,
        tran_date           DATE      NOT NULL,
        shift               VARCHAR(8) NOT NULL,
        sched_in_time       TIME      NULL,
        sched_out_time      TIME      NULL,
        actual_in           DATETIME  NULL,
        actual_out          DATETIME  NULL,
        work_dur_minutes    INT       NOT NULL DEFAULT 0,
        ot_minutes          INT       NOT NULL DEFAULT 0,
        break_minutes       INT       NOT NULL DEFAULT 0,
        total_dur_minutes   INT       NOT NULL DEFAULT 0,
        late_by_minutes     INT       NOT NULL DEFAULT 0,
        early_going_minutes INT       NOT NULL DEFAULT 0,
        status              VARCHAR(64) NOT NULL,
        punch_records       TEXT      NULL,
        created_at          DATETIME  NOT NULL DEFAULT CURRENT_TIMESTAMP,
        UNIQUE KEY uniq_eb_date (eb_id, tran_date)
    )
    """
)

_DELETE_DAILY_BASIC_SQL = text(
    "DELETE FROM daily_attendance_basic WHERE tran_date = :tran_date"
)

# Defensive migration: add break_minutes to a daily_attendance_basic that was
# created before this column existed. CREATE TABLE IF NOT EXISTS won't add it.
# IF NOT EXISTS on ADD COLUMN is MySQL 8.0.29+/MariaDB 10.0+; failures are
# swallowed by the caller (column already present, or older server).
_ADD_BREAK_COL_SQL = text(
    "ALTER TABLE daily_attendance_basic "
    "ADD COLUMN IF NOT EXISTS break_minutes INT NOT NULL DEFAULT 0 "
    "AFTER ot_minutes"
)

_INSERT_DAILY_BASIC_SQL = text(
    """
    INSERT INTO daily_attendance_basic
        (eb_id, emp_code, bio_id, dept_id, desig_id, tran_date, shift,
         sched_in_time, sched_out_time, actual_in, actual_out,
         work_dur_minutes, ot_minutes, break_minutes, total_dur_minutes,
         late_by_minutes, early_going_minutes, status, punch_records)
    VALUES
        (:eb_id, :emp_code, :bio_id, :dept_id, :desig_id, :tran_date, :shift,
         :sched_in_time, :sched_out_time, :actual_in, :actual_out,
         :work_dur_minutes, :ot_minutes, :break_minutes, :total_dur_minutes,
         :late_by_minutes, :early_going_minutes, :status, :punch_records)
    """
)


def _format_punch_records(
    relevant_punches: list[tuple],
    *,
    no_out_punch: bool,
    sched_out_str: str,
) -> str:
    """Format an employee's punches as a single comma-separated string,
    mirroring the eSSL 'Punch Records' column.

    Each entry is "HH:MM:direction(dev_<device_id>)". When there is no out
    punch, an "(SE)" sentinel record is appended at the scheduled out time.
    """
    parts: list[str] = []
    for p in relevant_punches:
        log_dt = p[5]  # log_date
        direction = (p[6] or "").lower() or "in"
        if isinstance(log_dt, datetime):
            ts = log_dt.strftime("%H:%M:%S")
        elif isinstance(log_dt, str):
            ts = log_dt[-8:] if len(log_dt) >= 8 else log_dt
        else:
            ts = str(log_dt)
        # device_id isn't carried in the tuple — keep the format compact.
        parts.append(f"{ts}:{direction}")
    if no_out_punch:
        # System-emitted out at scheduled end (matches eSSL "(SE)").
        parts.append(f"{sched_out_str}:out(SE)")
    return ",".join(parts)

FETCH_BPROCESS_PUNCHES_SQL = text(
    """
    SELECT b.eb_id,
           b.emp_code,
           b.bio_att_log_id,
           b.dept_id,
           b.desig_id,
           DATE(b.log_date)   AS punch_date,
           TIME(b.log_date)   AS punch_time,
           b.log_date         AS log_date,
           b.device_direction AS device_direction
    FROM bio_attendance_table b
    WHERE b.eb_id     IS NOT NULL
      AND b.dept_id   IS NOT NULL
      AND b.desig_id  IS NOT NULL
      AND b.device_id IS NOT NULL
      AND (
            DATE(b.log_date) = :tran_date
         OR (DATE(b.log_date) = DATE_ADD(:tran_date, INTERVAL 1 DAY)
             AND TIME(b.log_date) <= '08:00:00')
      )
    ORDER BY b.eb_id, b.log_date
    """
)


# Step 1 — prior night-shift OUT carryover.
# If yesterday's first punch was after 21:00, that employee is a night-shift
# worker, and their OUT happens on tran_date in [00:00..06:00]. Mark the LAST
# such punch as 'out' so it doesn't get treated as today's first IN.
_MARK_PRIOR_NIGHT_OUT_SQL = text(
    """
    UPDATE bio_attendance_table b
    JOIN (
        SELECT t.eb_id, MAX(t.log_date) AS edge_log
        FROM bio_attendance_table t
        JOIN (
            SELECT eb_id, MIN(log_date) AS first_log
            FROM bio_attendance_table
            WHERE eb_id IS NOT NULL
              AND DATE(log_date) = DATE_SUB(:tran_date, INTERVAL 1 DAY)
              AND (device_direction IS NULL OR device_direction <> 'out')
            GROUP BY eb_id
        ) y ON t.eb_id = y.eb_id
        WHERE DATE(t.log_date) = :tran_date
          AND TIME(t.log_date) <= '06:00:00'
          AND HOUR(y.first_log) > 21
        GROUP BY t.eb_id
    ) m ON b.eb_id = m.eb_id AND b.log_date = m.edge_log
    SET b.device_direction = 'out'
    WHERE DATE(b.log_date) = :tran_date
    """
)

# Step 2 — first IN of tran_date: earliest punch NOT already marked 'out'.
_MARK_FIRST_IN_SQL = text(
    """
    UPDATE bio_attendance_table b
    JOIN (
        SELECT eb_id, MIN(log_date) AS edge_log
        FROM bio_attendance_table
        WHERE eb_id IS NOT NULL
          AND DATE(log_date) = :tran_date
          AND (device_direction IS NULL OR device_direction <> 'out')
        GROUP BY eb_id
    ) m ON b.eb_id = m.eb_id AND b.log_date = m.edge_log
    SET b.device_direction = 'in'
    WHERE DATE(b.log_date) = :tran_date
    """
)

# Step 2b — rescue first IN when the device mislabels EVERY punch of an
# employee's day as 'out' (so Step 2 found no 'in' candidate and the day would be
# silently dropped). For each employee with NO 'in' punch on tran_date, mark
# their earliest punch that is NOT a carried-over night-shift exit as 'in'. The
# night-exit window ([00:00..06:00] for someone whose previous day started after
# 21:00) is excluded so a night worker's morning exit isn't mistaken for today's
# IN. Employees who already have any 'in' are untouched — normal days are
# unaffected; this only rescues the all-'out' case.
_MARK_RESCUE_FIRST_IN_SQL = text(
    """
    UPDATE bio_attendance_table b
    JOIN (
        SELECT t.eb_id, MIN(t.log_date) AS edge_log
        FROM bio_attendance_table t
        LEFT JOIN (
            SELECT eb_id, MIN(log_date) AS y_first
            FROM bio_attendance_table
            WHERE eb_id IS NOT NULL
              AND DATE(log_date) = DATE_SUB(:tran_date, INTERVAL 1 DAY)
            GROUP BY eb_id
        ) y ON t.eb_id = y.eb_id
        WHERE DATE(t.log_date) = :tran_date
          AND t.eb_id IS NOT NULL
          AND NOT (
                y.y_first IS NOT NULL
                AND HOUR(y.y_first) > 21
                AND TIME(t.log_date) <= '06:00:00'
          )
          AND t.eb_id NOT IN (
                SELECT eb_id FROM (
                    SELECT DISTINCT eb_id
                    FROM bio_attendance_table
                    WHERE DATE(log_date) = :tran_date
                      AND eb_id IS NOT NULL
                      AND LOWER(device_direction) = 'in'
                ) have_in
          )
        GROUP BY t.eb_id
    ) m ON b.eb_id = m.eb_id AND b.log_date = m.edge_log
    SET b.device_direction = 'in'
    WHERE DATE(b.log_date) = :tran_date
    """
)

# Step 3 — day-shift OUT: last punch on tran_date when the first IN's hour
# is <= 21. The first-IN row itself is excluded so a single-punch day stays
# as 'in' rather than being flipped to 'out'.
_MARK_LAST_OUT_DAY_SQL = text(
    """
    UPDATE bio_attendance_table b
    JOIN (
        SELECT p.eb_id, MAX(p.log_date) AS edge_log
        FROM bio_attendance_table p
        JOIN (
            SELECT eb_id, MIN(log_date) AS first_in
            FROM bio_attendance_table
            WHERE eb_id IS NOT NULL
              AND DATE(log_date) = :tran_date
              AND (device_direction IS NULL OR device_direction <> 'out')
            GROUP BY eb_id
        ) f ON p.eb_id = f.eb_id
        WHERE DATE(p.log_date) = :tran_date
          AND HOUR(f.first_in) <= 21
          AND p.log_date <> f.first_in
        GROUP BY p.eb_id
    ) m ON b.eb_id = m.eb_id AND b.log_date = m.edge_log
    SET b.device_direction = 'out'
    WHERE DATE(b.log_date) = :tran_date
    """
)

# Step 4 — night-shift OUT: for employees whose first IN on tran_date is
# after 21:00, the OUT is the LAST punch on tran_date+1 within [00:00..06:00].
_MARK_LAST_OUT_NIGHT_SQL = text(
    """
    UPDATE bio_attendance_table b
    JOIN (
        SELECT n.eb_id, MAX(n.log_date) AS edge_log
        FROM bio_attendance_table n
        JOIN (
            SELECT eb_id, MIN(log_date) AS first_in
            FROM bio_attendance_table
            WHERE eb_id IS NOT NULL
              AND DATE(log_date) = :tran_date
              AND (device_direction IS NULL OR device_direction <> 'out')
            GROUP BY eb_id
        ) f ON n.eb_id = f.eb_id
        WHERE DATE(n.log_date) = DATE_ADD(:tran_date, INTERVAL 1 DAY)
          AND TIME(n.log_date) <= '06:00:00'
          AND HOUR(f.first_in) > 21
        GROUP BY n.eb_id
    ) m ON b.eb_id = m.eb_id AND b.log_date = m.edge_log
    SET b.device_direction = 'out'
    WHERE DATE(b.log_date) = DATE_ADD(:tran_date, INTERVAL 1 DAY)
    """
)


def _split_worked_break_secs(
    punch_secs: list[int], first_sec: int, last_sec: int
) -> tuple[int, int]:
    """Split the span [first_sec, last_sec] into worked vs break seconds.

    `first_sec` is the marked first-IN and `last_sec` the marked last-OUT
    (both reliable). Punches strictly between them come in (OUT, IN) break
    pairs — the employee punches OUT to leave, then IN to return — so each
    such pair's gap is break time and worked = span − breaks.

    A trailing unpaired intermediate punch (odd count → a missing break-out or
    break-in) is ignored: no break is subtracted for it, degrading to the
    plain-span result rather than guessing.

    Returns ``(worked_secs, break_secs)``.
    """
    total_secs = max(0, last_sec - first_sec)
    mid = sorted(s for s in punch_secs if first_sec < s < last_sec)
    break_secs = 0
    j = 0
    while j + 1 < len(mid):  # pair (OUT, IN); drop a trailing unpaired punch
        break_secs += mid[j + 1] - mid[j]
        j += 2
    break_secs = min(break_secs, total_secs)
    return total_secs - break_secs, break_secs


def _process_bprocess_day(
    db: Session,
    *,
    tran_date: str,
    is_off_day: bool,
) -> int:
    """Insert eSSL-style daily rows for each employee with punches on tran_date.

    Writes:
      - daily_attendance_process_table : one spell row (A/B/C/GS).
      - daily_attendance_basic         : one detailed row mirroring the eSSL
                                         Daily Attendance Detailed Report.

    Punches from tran_date+1 [00:00..06:00] feed the night-shift (C) OUT.
    """
    # Marking pipeline (order matters):
    #   1. Carry over yesterday's night-shift OUT into today's [00:00..06:00].
    #   2. Mark today's first IN as the earliest punch NOT already 'out'.
    #   3. Mark today's day-shift OUT (last punch when first IN <= 21).
    #   4. Mark night-shift OUT on tran_date+1 (when first IN > 21).
    db.execute(_MARK_PRIOR_NIGHT_OUT_SQL, {"tran_date": tran_date})
    db.execute(_MARK_FIRST_IN_SQL, {"tran_date": tran_date})
    # Rescue: when the device labelled every punch 'out', mark the earliest
    # non-night-exit punch as the IN so the day still produces attendance.
    db.execute(_MARK_RESCUE_FIRST_IN_SQL, {"tran_date": tran_date})
    db.execute(_MARK_LAST_OUT_DAY_SQL, {"tran_date": tran_date})
    db.execute(_MARK_LAST_OUT_NIGHT_SQL, {"tran_date": tran_date})

    # Ensure target table exists; wipe rows for this tran_date for re-run.
    db.execute(_CREATE_DAILY_BASIC_SQL)
    # Best-effort: add break_minutes to tables created before it existed.
    try:
        db.execute(_ADD_BREAK_COL_SQL)
    except Exception:
        db.rollback()
    db.execute(_DELETE_DAILY_BASIC_SQL, {"tran_date": tran_date})

    rows = db.execute(
        FETCH_BPROCESS_PUNCHES_SQL, {"tran_date": tran_date},
    ).fetchall()

    by_emp: dict[int, list[tuple]] = {}
    emp_codes: dict[int, str | None] = {}
    for r in rows:
        m = r._mapping
        if m["eb_id"] not in emp_codes:
            emp_codes[m["eb_id"]] = m["emp_code"]
        by_emp.setdefault(m["eb_id"], []).append((
            str(m["punch_date"]), m["punch_time"], m["bio_att_log_id"],
            m["dept_id"], m["desig_id"], m["log_date"],
            m["device_direction"],
        ))

    inserted = 0
    for eb_id, punches in by_emp.items():
        emp_code = emp_codes.get(eb_id)
        day0 = [p for p in punches if p[0] == tran_date]
        day1 = [p for p in punches if p[0] != tran_date]
        if not day0:
            continue

        # Skip leading day0 rows already marked 'out' — those are the prior
        # night shift's exit punches that landed in today's date.
        first_idx = 0
        while (
            first_idx < len(day0)
            and (day0[first_idx][6] or "").lower() == "out"
        ):
            first_idx += 1
        if first_idx >= len(day0):
            continue
        day0 = day0[first_idx:]

        first_pdate, first_time, first_bio, first_dept, first_desig, first_log_date, _ = day0[0]
        first_sec = _to_seconds(first_time)
        first_h   = first_sec / 3600.0

        # ---- Step 1: pick OUT punch ----
        # Priority: if next day's first record is 'out' AND the resulting
        # span (first_in -> that punch) is < 17h, that punch is today's OUT.
        # Cross-midnight cases (any first_h) are caught here.
        no_out_punch = False
        last_sec: int | None = None
        last_log_date = None
        crosses_midnight = False

        if day1 and (day1[0][6] or "").lower() == "out":
            cand_sec = _to_seconds(day1[0][1]) + 24 * 3600
            if cand_sec - first_sec < 17 * 3600:
                last_sec = cand_sec
                last_log_date = day1[0][5]
                crosses_midnight = True

        # ---- Step 2: classify shift ----
        # Cross-midnight overrides band detection → always Shift C.
        if crosses_midnight:
            shift = "C"
        else:
            shift = _bprocess_shift_for(first_h)
            if shift is None:
                continue  # No matching shift band → treat as NS, skip.

        sched_in_str, sched_out_str = _BPROCESS_SPELL_TIMES[shift]
        sched_in_sec = _to_seconds(_parse_hms(sched_in_str))
        sched_out_sec = _to_seconds(_parse_hms(sched_out_str))
        if shift == "C":  # OUT is on next day → 06:00 + 24h.
            sched_out_sec += 24 * 3600

        # ---- Step 3: fall back to shift-specific OUT when priority didn't hit. ----
        if last_sec is None:
            if shift == "C":
                if day1:
                    _, last_time, _, _, _, last_log_date, _ = day1[-1]
                    last_sec = _to_seconds(last_time) + 24 * 3600
                elif len(day0) > 1:
                    _, last_time, _, _, _, last_log_date, _ = day0[-1]
                    last_sec = _to_seconds(last_time)
                else:
                    # No OUT punch found — synthesize at scheduled out time.
                    no_out_punch = True
                    last_sec = sched_out_sec
                    last_log_date = None
            else:
                if len(day0) > 1:
                    _, last_time, _, _, _, last_log_date, _ = day0[-1]
                    last_sec = _to_seconds(last_time)
                else:
                    no_out_punch = True
                    last_sec = sched_out_sec
                    last_log_date = None

        total_secs = max(0, last_sec - first_sec)

        # ---- Break-aware worked time ----
        # Punches between the first-IN and last-OUT come in (OUT, IN) break
        # pairs; subtract each gap as break time. See _split_worked_break_secs.
        # Example: 05:54 IN, 13:55 OUT, 18:03 IN, 21:54 OUT → break 13:55→18:03,
        # worked = 16h − 4.13h = 11.85h (8h work + 3.85h OT).
        punch_secs = [_to_seconds(p[1]) for p in day0]
        if crosses_midnight or shift == "C":
            punch_secs.extend(_to_seconds(p[1]) + 24 * 3600 for p in day1)
        worked_secs, break_secs = _split_worked_break_secs(
            punch_secs, first_sec, last_sec
        )

        # eSSL Work/OT split — applied to the break-excluded worked time.
        if no_out_punch:
            # No OUT punch — use sched_out as boundary; whole span is Work
            # (no OT credit because there's no actual out-time to verify).
            work_secs = worked_secs
            ot_secs   = 0
        else:
            work_secs = min(_REGULAR_WORK_SECONDS, worked_secs)
            ot_secs   = max(0, worked_secs - _REGULAR_WORK_SECONDS)

        late_by_secs       = max(0, first_sec - sched_in_sec)
        early_going_secs   = (
            0 if no_out_punch else max(0, sched_out_sec - last_sec)
        )

        if no_out_punch:
            status = "Present (No OutPunch)"
        else:
            status = "Off Day Present" if is_off_day else "Present"

        # ---- daily_attendance_process_table (one spell row) ----
        att_type      = "O" if is_off_day else "R"
        working_hours = round(work_secs / 3600.0, 2)
        ot_hours      = round(ot_secs / 3600.0, 2)
        time_duration = round(working_hours + ot_hours, 2)
        spell_start, spell_end = _BPROCESS_SPELL_TIMES[shift]
        db.execute(
            INSERT_SPELL_ROW_SQL,
            {
                "eb_id":           int(eb_id),
                "bio_id":          int(first_bio) if first_bio is not None else None,
                "dept_id":         int(first_dept) if first_dept is not None else None,
                "desig_id":        int(first_desig) if first_desig is not None else None,
                "tran_date":       tran_date,
                "spell_name":      shift,
                "attendance_type": att_type,
                "check_in":        first_log_date,
                "check_out":       last_log_date,
                "time_duration":   time_duration,
                "working_hours":   working_hours,
                "ot_hours":        ot_hours,
                "spell_start":     spell_start,
                "spell_end":       spell_end,
                "spell_hours":     SPELL_HOURS,
            },
        )

        # ---- daily_attendance_basic (eSSL detailed-report row) ----
        # Build the punch_records string from today's relevant punches +
        # any next-day punch we used for the OUT (cross-midnight / shift C).
        relevant_punches = list(day0)
        if crosses_midnight or shift == "C":
            relevant_punches.extend(day1)
        punch_records_str = _format_punch_records(
            relevant_punches,
            no_out_punch=no_out_punch,
            sched_out_str=sched_out_str,
        )

        db.execute(
            _INSERT_DAILY_BASIC_SQL,
            {
                "eb_id":               int(eb_id),
                "emp_code":            emp_code,
                "bio_id":              int(first_bio) if first_bio is not None else None,
                "dept_id":             int(first_dept) if first_dept is not None else None,
                "desig_id":            int(first_desig) if first_desig is not None else None,
                "tran_date":           tran_date,
                "shift":               shift,
                "sched_in_time":       sched_in_str,
                "sched_out_time":      sched_out_str,
                "actual_in":           first_log_date,
                "actual_out":          last_log_date,
                "work_dur_minutes":    work_secs // 60,
                "ot_minutes":          ot_secs // 60,
                "break_minutes":       break_secs // 60,
                "total_dur_minutes":   total_secs // 60,
                "late_by_minutes":     late_by_secs // 60,
                "early_going_minutes": early_going_secs // 60,
                "status":              status,
                "punch_records":       punch_records_str,
            },
        )
        inserted += 1

    print(f"[bio_att_bprocess]   -> rows inserted = {inserted}", flush=True)
    return inserted


@router.post("/bio_att_bprocess")
async def bio_att_bprocess(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Same orchestration as /bio_att_etrack_process — resolve eb_id /
    dept_id / desig_id on bio_attendance_table, then build daily rows — but
    using the Bprocess rules (Rules #1/#2/#3 above, including cross-midnight
    C/C1 spell)."""
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        body: dict = {}
        try:
            body = await request.json()
        except Exception:
            pass
        qp = request.query_params

        tran_date_raw = body.get("tran_date") or qp.get("tran_date")
        if not tran_date_raw:
            raise HTTPException(status_code=400, detail="tran_date is required")
        try:
            datetime.strptime(tran_date_raw, "%Y-%m-%d")
            tran_date: str = tran_date_raw
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid tran_date {tran_date_raw!r}, expected YYYY-MM-DD",
            )

        branch_id_raw = body.get("branch_id") or qp.get("branch_id")
        if branch_id_raw in (None, ""):
            raise HTTPException(status_code=400, detail="branch_id is required")
        try:
            branch_id = int(branch_id_raw)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="branch_id must be an integer")

        # Resolve + process with Bprocess rules (shared with the pipeline).
        core = bprocess_core(db, tran_date)

        return {
            "status": "ok",
            "tran_date": tran_date,
            "branch_id": branch_id,
            "is_off_day": core["is_off_day"],
            "resolve": core["resolve"],
            "process": {
                "total_inserted": core["inserted"],
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


@router.post("/bio_att_b_atten")
async def bio_att_b_atten(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Project rows from `daily_attendance_basic` into
    `daily_attendance_process_table` for the given tran_date.

    Use this when the basic table has already been populated/edited and you
    want spell rows refreshed without re-running the full bprocess pipeline.

    Required body params:
        tran_date : YYYY-MM-DD
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        body: dict = {}
        try:
            body = await request.json()
        except Exception:
            pass
        qp = request.query_params

        tran_date_raw = body.get("tran_date") or qp.get("tran_date")
        if not tran_date_raw:
            raise HTTPException(status_code=400, detail="tran_date is required")
        try:
            datetime.strptime(tran_date_raw, "%Y-%m-%d")
            tran_date: str = tran_date_raw
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid tran_date {tran_date_raw!r}, expected YYYY-MM-DD",
            )

        # Project daily_attendance_basic -> process table (shared with pipeline).
        core = b_atten_core(db, tran_date)

        return {
            "status": "ok",
            "tran_date": tran_date,
            "is_off_day": core["is_off_day"],
            "basic_rows": core["basic_rows"],
            "inserted": core["inserted"],
        }

    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


def _parse_etrack_proc_params(body: dict, qp) -> tuple[str, int]:
    """Validate tran_date (YYYY-MM-DD) and branch_id shared by the Etrack
    Process endpoints. Raises HTTPException on bad input."""
    tran_date_raw = body.get("tran_date") or qp.get("tran_date")
    if not tran_date_raw:
        raise HTTPException(status_code=400, detail="tran_date is required")
    try:
        datetime.strptime(tran_date_raw, "%Y-%m-%d")
        tran_date = tran_date_raw
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid tran_date {tran_date_raw!r}, expected YYYY-MM-DD",
        )

    branch_id_raw = body.get("branch_id") or qp.get("branch_id")
    if branch_id_raw in (None, ""):
        raise HTTPException(status_code=400, detail="branch_id is required")
    try:
        branch_id = int(branch_id_raw)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="branch_id must be an integer")
    return tran_date, branch_id


def _etrack_resolve_and_process(db: Session, tran_date: str) -> dict:
    """Shared Etrack Process core. Resolves eb_id / dept_id / desig_id for
    bio_attendance rows where eb_id IS NULL (via tbl_master_bio_link_mst, then
    last daily_attendance, then hrms_ed_official_details), then deletes and
    rebuilds the daily_attendance_process_table spell rows for tran_date.

    Returns {"resolve": {...}, "is_off_day": bool, "inserted": int}.
    """
    # ── Step 1: distinct emp_code -> eb_id from link master ─────────────
    unresolved_rows = db.execute(_ETRACK_PROC_UNRESOLVED_SQL).fetchall()

    resolve_result = {
        "resolved": 0,
        "updated": 0,
        "from_daily_attendance": 0,
        "fallback_official": 0,
        "no_source": 0,
    }

    if unresolved_rows:
        emp_eb_map: dict[str, int] = {
            str(r.emp_code): int(r.eb_id)
            for r in unresolved_rows
            if r.eb_id is not None
        }

        if emp_eb_map:
            unique_eb_ids = list(set(emp_eb_map.values()))

            # ── Step 2: last daily_attendance per eb_id ──────────────────
            da_rows = db.execute(
                _ETRACK_PROC_LAST_DAILY_ATT_SQL,
                {"eb_ids": tuple(unique_eb_ids)},
            ).fetchall()
            da_map: dict[int, tuple] = {
                int(r.eb_id): (r.worked_department_id, r.worked_designation_id)
                for r in da_rows
                if r.worked_department_id is not None
            }

            # ── Step 3: fallback from hrms_ed_official_details ───────────
            missing_eb_ids = [eid for eid in unique_eb_ids if eid not in da_map]
            official_map: dict[int, tuple] = {}
            if missing_eb_ids:
                off_rows = db.execute(
                    _ETRACK_PROC_OFFICIAL_SQL,
                    {"eb_ids": tuple(missing_eb_ids)},
                ).fetchall()
                official_map = {
                    int(r.eb_id): (r.dept_id, r.desig_id)
                    for r in off_rows
                }

            # ── Step 4: UPDATE bio_attendance_table ──────────────────────
            updated = 0
            fallback_official = 0
            no_source = 0

            for emp_code, eb_id in emp_eb_map.items():
                if eb_id in da_map:
                    dept_id, desig_id = da_map[eb_id]
                elif eb_id in official_map:
                    dept_id, desig_id = official_map[eb_id]
                    fallback_official += 1
                else:
                    no_source += 1
                    continue

                res = db.execute(
                    _ETRACK_PROC_UPDATE_SQL,
                    {
                        "eb_id": eb_id,
                        "dept_id": dept_id,
                        "desig_id": desig_id,
                        "emp_code": emp_code,
                    },
                )
                updated += int(res.rowcount or 0)

            db.commit()

            resolve_result = {
                "resolved": len(emp_eb_map),
                "updated": updated,
                "from_daily_attendance": len(da_map),
                "fallback_official": fallback_official,
                "no_source": no_source,
            }

    # ── Step 5: process into daily_attendance_process_table ─────────────
    is_off_row = db.execute(IS_OFF_DAY_SQL, {"tran_date": tran_date}).fetchone()
    is_off_day = bool(is_off_row and int(is_off_row.cnt) > 0)

    # Delete existing rows for the date (re-run safe).
    db.execute(DELETE_DAY_ROWS_SQL, {"tran_date": tran_date})
    db.commit()

    inserted = _process_etrack_day(
        db, tran_date=tran_date, is_off_day=is_off_day,
    )
    db.commit()

    return {
        "resolve": resolve_result,
        "is_off_day": is_off_day,
        "inserted": inserted,
    }


def etrack_process_d_prepare(db: Session) -> dict:
    """Table-global Etrack pre-steps (0/0b/0c) that do NOT depend on any single
    date. Safe — and cheaper — to run once per pipeline pass rather than once per
    date (each is a full-table UPDATE on bio_attendance_table):

      0.  emp_code      <- tbl_master_bio_link_mst.master_data
      0b. bio_att_log_id <- generated 500000+ sequence where blank (NULL or 0).
      0c. device_id      <- 22 when device_direction='in', else 14.

    Returns the three pre-step row counts.
    """
    # ── Step 0: back-fill emp_code from link master (bio_id -> emp_code) ─
    empcode_res = db.execute(_ETRACK_PROC_EMPCODE_UPDATE_SQL)
    emp_code_updated = int(empcode_res.rowcount or 0)
    db.commit()

    # ── Step 0b: fill blank bio_att_log_id with a generated sequence ────
    logid_max_row = db.execute(
        text("SELECT MAX(bio_att_log_id) AS max_id FROM bio_attendance_table")
    ).mappings().first()
    cur_max_log_id = (
        int(logid_max_row["max_id"])
        if logid_max_row and logid_max_row["max_id"] is not None
        else 0
    )
    logid_base = 500000 if cur_max_log_id < 500000 else cur_max_log_id + 1
    db.execute(_ETRACK_PROC_LOGID_INIT_SQL, {"base": logid_base})
    logid_res = db.execute(_ETRACK_PROC_LOGID_FILL_SQL)
    bio_att_log_id_filled = int(logid_res.rowcount or 0)
    db.commit()

    # ── Step 0c: set device_id from direction ('in'->22, else 14) ───────
    device_id_res = db.execute(_ETRACK_PROC_DEVICE_ID_SQL)
    device_id_updated = int(device_id_res.rowcount or 0)
    db.commit()

    return {
        "emp_code_updated": emp_code_updated,
        "bio_att_log_id_filled": bio_att_log_id_filled,
        "device_id_updated": device_id_updated,
    }


def etrack_process_d_core(
    db: Session, tran_date: str, *, run_prepare: bool = True
) -> dict:
    """Etrack Process (D) core. Runs the three pre-steps on bio_attendance_table
    then the shared resolve + process. Shared by the /bio_att_etrack_process_d
    route and the automated pipeline (bio_att_auto_pipeline).

    The pre-steps are table-global (date-independent). Pass run_prepare=False when
    the caller has already run etrack_process_d_prepare once for the whole pass
    (the automated pipeline does this) so they aren't repeated per date.

    Returns the resolve/process dict plus the three pre-step row counts.
    """
    prep = (
        etrack_process_d_prepare(db)
        if run_prepare
        else {
            "emp_code_updated": 0,
            "bio_att_log_id_filled": 0,
            "device_id_updated": 0,
        }
    )

    # Resolve eb_id / dept_id / desig_id, then build the day's spell rows.
    core = _etrack_resolve_and_process(db, tran_date)
    return {**prep, **core}


def bprocess_core(db: Session, tran_date: str) -> dict:
    """Bprocess core. Same eb_id/dept_id/desig_id resolve as Etrack Process, but
    builds the day's spell rows with the Bprocess rules (cross-midnight C/C1) and
    writes daily_attendance_basic via _process_bprocess_day. Shared by the
    /bio_att_bprocess route and the automated pipeline.

    Returns {"resolve": {...}, "is_off_day": bool, "inserted": int}.
    """
    unresolved_rows = db.execute(_ETRACK_PROC_UNRESOLVED_SQL).fetchall()

    resolve_result = {
        "resolved": 0,
        "updated": 0,
        "from_daily_attendance": 0,
        "fallback_official": 0,
        "no_source": 0,
    }

    if unresolved_rows:
        emp_eb_map: dict[str, int] = {
            str(r.emp_code): int(r.eb_id)
            for r in unresolved_rows
            if r.eb_id is not None
        }

        if emp_eb_map:
            unique_eb_ids = list(set(emp_eb_map.values()))

            da_rows = db.execute(
                _ETRACK_PROC_LAST_DAILY_ATT_SQL,
                {"eb_ids": tuple(unique_eb_ids)},
            ).fetchall()
            da_map: dict[int, tuple] = {
                int(r.eb_id): (r.worked_department_id, r.worked_designation_id)
                for r in da_rows
                if r.worked_department_id is not None
            }

            missing_eb_ids = [eid for eid in unique_eb_ids if eid not in da_map]
            official_map: dict[int, tuple] = {}
            if missing_eb_ids:
                off_rows = db.execute(
                    _ETRACK_PROC_OFFICIAL_SQL,
                    {"eb_ids": tuple(missing_eb_ids)},
                ).fetchall()
                official_map = {
                    int(r.eb_id): (r.dept_id, r.desig_id)
                    for r in off_rows
                }

            updated = 0
            fallback_official = 0
            no_source = 0

            for emp_code, eb_id in emp_eb_map.items():
                if eb_id in da_map:
                    dept_id, desig_id = da_map[eb_id]
                elif eb_id in official_map:
                    dept_id, desig_id = official_map[eb_id]
                    fallback_official += 1
                else:
                    no_source += 1
                    continue

                res = db.execute(
                    _ETRACK_PROC_UPDATE_SQL,
                    {
                        "eb_id": eb_id,
                        "dept_id": dept_id,
                        "desig_id": desig_id,
                        "emp_code": emp_code,
                    },
                )
                updated += int(res.rowcount or 0)

            db.commit()

            resolve_result = {
                "resolved": len(emp_eb_map),
                "updated": updated,
                "from_daily_attendance": len(da_map),
                "fallback_official": fallback_official,
                "no_source": no_source,
            }

    is_off_row = db.execute(IS_OFF_DAY_SQL, {"tran_date": tran_date}).fetchone()
    is_off_day = bool(is_off_row and int(is_off_row.cnt) > 0)

    db.execute(DELETE_DAY_ROWS_SQL, {"tran_date": tran_date})
    db.commit()

    inserted = _process_bprocess_day(
        db, tran_date=tran_date, is_off_day=is_off_day,
    )
    db.commit()

    return {
        "resolve": resolve_result,
        "is_off_day": is_off_day,
        "inserted": inserted,
    }


def b_atten_core(db: Session, tran_date: str) -> dict:
    """B Atten core. Projects rows from daily_attendance_basic into
    daily_attendance_process_table for tran_date. Shared by the /bio_att_b_atten
    route and the automated pipeline.

    Returns {"is_off_day": bool, "basic_rows": int, "inserted": int}.
    """
    # Off-day flag drives attendance_type R vs O.
    off_row = db.execute(IS_OFF_DAY_SQL, {"tran_date": tran_date}).first()
    is_off_day = bool(off_row[0]) if off_row else False

    # Wipe existing spell rows for the date so this is idempotent.
    db.execute(DELETE_DAY_ROWS_SQL, {"tran_date": tran_date})

    basic_rows = db.execute(
        text(
            """
            SELECT eb_id, bio_id, dept_id, desig_id, shift,
                   actual_in, actual_out,
                   work_dur_minutes, ot_minutes, total_dur_minutes
            FROM daily_attendance_basic
            WHERE tran_date = :tran_date
            """
        ),
        {"tran_date": tran_date},
    ).fetchall()

    inserted = 0
    for r in basic_rows:
        m = r._mapping
        shift = m["shift"]
        actual_in = m["actual_in"]
        intime_h = (
            actual_in.hour + actual_in.minute / 60.0 + actual_in.second / 3600.0
            if isinstance(actual_in, datetime) else 0.0
        )
        work_min = int(m["work_dur_minutes"] or 0)
        ot_min   = int(m["ot_minutes"] or 0)

        common_base = {
            "eb_id":     int(m["eb_id"]),
            "bio_id":    int(m["bio_id"]) if m["bio_id"] is not None else None,
            "dept_id":   int(m["dept_id"]) if m["dept_id"] is not None else None,
            "desig_id":  int(m["desig_id"]) if m["desig_id"] is not None else None,
            "tran_date": tran_date,
            "check_in":   actual_in,
            "check_out":  m["actual_out"],
            "spell_hours": SPELL_HOURS,
        }

        # ── R (working-hours) row ──
        r_bucket = _minutes_bucket(work_min)
        if r_bucket >= 0:
            spell = _spell_label_for_b_atten(shift, "R", r_bucket, intime_h)
            spell_start, spell_end = _BPROCESS_SPELL_TIMES.get(
                spell, ("00:00:00", "00:00:00"),
            )
            db.execute(
                INSERT_SPELL_ROW_SQL,
                {
                    **common_base,
                    "spell_name":      spell,
                    "attendance_type": "O" if is_off_day else "R",
                    "time_duration":   r_bucket,
                    "working_hours":   r_bucket,
                    "ot_hours":        0,
                    "spell_start":     spell_start,
                    "spell_end":       spell_end,
                },
            )
            inserted += 1

        # ── O (overtime) row ──
        o_bucket = _minutes_bucket(ot_min)
        if o_bucket >= 0:
            spell = _spell_label_for_b_atten(shift, "O", o_bucket, intime_h)
            spell_start, spell_end = _BPROCESS_SPELL_TIMES.get(
                spell, ("00:00:00", "00:00:00"),
            )
            db.execute(
                INSERT_SPELL_ROW_SQL,
                {
                    **common_base,
                    "spell_name":      spell,
                    "attendance_type": "O",
                    "time_duration":   o_bucket,
                    "working_hours":   o_bucket,
                    "ot_hours":        0,
                    "spell_start":     spell_start,
                    "spell_end":       spell_end,
                },
            )
            inserted += 1

    db.commit()

    return {
        "is_off_day": is_off_day,
        "basic_rows": len(basic_rows),
        "inserted": inserted,
    }


@router.post("/bio_att_etrack_process")
async def bio_att_etrack_process(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Resolve eb_id / dept_id / desig_id for bio_attendance rows where eb_id IS NULL,
    then process them into daily_attendance_process_table for the given date.

    Steps per unresolved emp_code:
      1. Look up eb_id from tbl_master_bio_link_mst (match_type='E').
      2. Check last record in daily_attendance for that eb_id:
           worked_department_id -> dept_id, worked_designation_id -> desig_id.
      3. If no daily_attendance record found, fall back to hrms_ed_official_details:
           sub_dept_id -> dept_id, designation_id -> desig_id.
      4. UPDATE bio_attendance_table for that emp_code (eb_id IS NULL rows only).
      5. Delete existing daily_attendance_process_table rows for tran_date,
         then run Spell A + Spell B processing into daily_attendance_process_table.

    Required body params:
        tran_date  : YYYY-MM-DD
        branch_id  : int
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        body: dict = {}
        try:
            body = await request.json()
        except Exception:
            pass
        qp = request.query_params

        tran_date, branch_id = _parse_etrack_proc_params(body, qp)

        # Resolve eb_id / dept_id / desig_id, then build the day's spell rows.
        core = _etrack_resolve_and_process(db, tran_date)

        return {
            "status": "ok",
            "tran_date": tran_date,
            "branch_id": branch_id,
            "is_off_day": core["is_off_day"],
            "resolve": core["resolve"],
            "process": {
                "total_inserted": core["inserted"],
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


@router.post("/bio_att_etrack_process_d")
async def bio_att_etrack_process_d(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Etrack Process (D). Runs three pre-steps on bio_attendance_table, then
    the same resolve + process as /bio_att_etrack_process:

      0.  emp_code      <- tbl_master_bio_link_mst.master_data
                            (match bio_id = bio_dev_id, match_type='E').
      0b. bio_att_log_id <- generated 500000+ sequence where blank (NULL or 0).
      0c. device_id      <- 22 when device_direction='in', else 14.

    These pre-steps run only here, NOT in /bio_att_etrack_process.

    Required body params:
        tran_date  : YYYY-MM-DD
        branch_id  : int
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        body: dict = {}
        try:
            body = await request.json()
        except Exception:
            pass
        qp = request.query_params

        tran_date, branch_id = _parse_etrack_proc_params(body, qp)

        # Pre-steps + resolve + process (shared with the automated pipeline).
        core = etrack_process_d_core(db, tran_date)

        return {
            "status": "ok",
            "tran_date": tran_date,
            "branch_id": branch_id,
            "is_off_day": core["is_off_day"],
            "emp_code_updated": core["emp_code_updated"],
            "bio_att_log_id_filled": core["bio_att_log_id_filled"],
            "device_id_updated": core["device_id_updated"],
            "resolve": core["resolve"],
            "process": {
                "total_inserted": core["inserted"],
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


# =============================================================================
# Wages Register (attwgs) — pivoted shift + wages report
# =============================================================================
#
# Per (employee, attendance_date):
#   shift_letter  = 'A' if check_in hour in [5..12], 'B' if [13..20]
#   shift_code    = shift_letter             when working_hours == 8
#                   shift_letter + '1'       otherwise
#   rate          = latest employee_rate_table.rate where rate_date <= attendance_date
#   wages         = (rate / 8) * (working_hours + ot_hours)

_WAGES_REGISTER_SQL = text(
    """
                    SELECT
        d.eb_id                                                          AS eb_id,
        o.emp_code                                                       AS emp_code,
        TRIM(CONCAT_WS(' ', p.first_name,
                            IFNULL(p.middle_name, ''),
                            IFNULL(p.last_name,  ''))) AS emp_name,
        sdm.sub_dept_desc                                                    AS department,
        dm.desig                                                     AS designation,
        d.attendance_date                                                AS attendance_date,
        d.spell_name                                                     AS spell_name,
        d.attendance_type                                                AS attendance_type,
        d.check_in                                                       AS check_in,
        d.Working_hours                                                  AS working_hours,
        d.Ot_hours                                                       AS ot_hours,
        (
            SELECT er.rate
            FROM employee_rate_table er
            WHERE er.eb_id = d.eb_id
              AND er.rate_date <= d.attendance_date
            ORDER BY er.rate_date DESC
            LIMIT 1
        )                                                                AS rate
    FROM daily_attendance_process_table d
    LEFT JOIN hrms_ed_official_details o ON o.eb_id = d.eb_id
    LEFT JOIN hrms_ed_personal_details p ON p.eb_id = d.eb_id
	left join sub_dept_mst sdm on o.sub_dept_id =sdm.sub_dept_id 
    left join designation_mst dm on dm.designation_id =o.designation_id 
	WHERE d.attendance_date BETWEEN :from_date AND :to_date
      AND (:branch_id = 0 OR o.branch_id = :branch_id)
      AND (:emp_code = '' OR o.emp_code = :emp_code)
    ORDER BY CAST(o.emp_code AS UNSIGNED), o.emp_code, d.attendance_date
    """
)
 
def _shift_letter_from_check_in(check_in) -> str:
    """A if check_in hour in [5..12], B if in [13..20], else ''."""
    if check_in is None:
        return ""
    if isinstance(check_in, datetime):
        hour = check_in.hour
    elif isinstance(check_in, dt_time):
        hour = check_in.hour
    elif isinstance(check_in, timedelta):
        hour = int(check_in.total_seconds() // 3600) % 24
    else:
        try:
            hour = int(str(check_in).split(":")[0])
        except Exception:
            return ""
    if 5 <= hour <= 12:
        return "A"
    if 13 <= hour <= 20:
        return "B"
    return ""


@router.get("/wages_register")
async def wages_register(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Pivoted wages register: shift code + per-day wages per employee."""
    qp = request.query_params
    co_id = qp.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    from_raw = qp.get("from_date")
    to_raw = qp.get("to_date")
    if not from_raw or not to_raw:
        raise HTTPException(status_code=400, detail="from_date and to_date are required")
    try:
        from_date = datetime.strptime(from_raw, "%Y-%m-%d").date()
        to_date = datetime.strptime(to_raw, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="dates must be YYYY-MM-DD")
    if from_date > to_date:
        raise HTTPException(status_code=400, detail="from_date must be <= to_date")

    branch_raw = qp.get("branch_id") or "0"
    try:
        branch_id = int(branch_raw)
    except ValueError:
        raise HTTPException(status_code=400, detail="branch_id must be an integer")

    emp_code = (qp.get("emp_code") or "").strip()

    try:
        rows = db.execute(
            _WAGES_REGISTER_SQL,
            {
                "from_date": from_date,
                "to_date": to_date,
                "branch_id": branch_id,
                "emp_code": emp_code,
            },
        ).fetchall()

        columns: list[dict] = []
        d = from_date
        while d <= to_date:
            columns.append({
                "key": d.isoformat(),
                "label": d.strftime("%a, %b %d"),
            })
            d += timedelta(days=1)

        emp_meta:      dict[int, dict] = {}
        emp_shifts:    dict[int, dict[str, str]] = {}
        emp_ot_shifts: dict[int, dict[str, str]] = {}
        emp_wages:     dict[int, dict[str, float]] = {}
        emp_whrs_by:   dict[int, dict[str, float]] = {}
        emp_ot_by:     dict[int, dict[str, float]] = {}
        emp_whrs:      dict[int, float] = {}
        emp_ot:        dict[int, float] = {}
        emp_total:     dict[int, float] = {}
        emp_rate:      dict[int, float] = {}
        emp_p_days:    dict[int, set[str]] = {}
        emp_ot_days:   dict[int, set[str]] = {}

        for r in rows:
            m = r._mapping
            eb_id = m["eb_id"]
            if eb_id is None:
                continue
            if eb_id not in emp_meta:
                emp_meta[eb_id] = {
                    "eb_id": int(eb_id),
                    "emp_code": m["emp_code"] or "",
                    "emp_name": m["emp_name"] or "",
                    "department": m["department"] or "",
                    "designation": m["designation"] or "",
                }
                emp_shifts[eb_id]    = {}
                emp_ot_shifts[eb_id] = {}
                emp_wages[eb_id]     = {}
                emp_whrs_by[eb_id]   = {}
                emp_ot_by[eb_id]     = {}
                emp_whrs[eb_id]      = 0.0
                emp_ot[eb_id]        = 0.0
                emp_total[eb_id]     = 0.0
                emp_rate[eb_id]      = float(m["rate"] or 0)
                emp_p_days[eb_id]    = set()
                emp_ot_days[eb_id]   = set()

            att_date = m["attendance_date"]
            if att_date is None:
                continue
            if isinstance(att_date, datetime):
                att_date = att_date.date()
            key = att_date.isoformat()

            wh = float(m["working_hours"] or 0)
            ot = float(m["ot_hours"] or 0)
            rate = float(m["rate"] or 0)
            att_type = (m["attendance_type"] or "").upper()

            shift_from_check = _shift_letter_from_check_in(m["check_in"])
            shift_letter = shift_from_check or (m["spell_name"] or "")
            if not shift_letter:
                continue

            # Route hours to the Shift vs OT cell by attendance_type, not by
            # which numeric column holds the value. _process_b_atten stores OT
            # rows with working_hours > 0 and ot_hours = 0 but attendance_type
            # = 'O' — those must land in the OT cell, not the Shift cell.
            #
            # The "1" suffix marks a partial bucket (hours != 8).
            if att_type == "O":
                hrs = wh + ot
                if hrs > 0:
                    code = shift_letter if hrs == 8 else f"{shift_letter}1"
                    existing = emp_ot_shifts[eb_id].get(key, "")
                    emp_ot_shifts[eb_id][key] = (
                        f"{existing} {code}".strip() if existing else code
                    )
                    emp_ot_by[eb_id][key] = round(
                        emp_ot_by[eb_id].get(key, 0.0) + hrs, 2
                    )
                    emp_ot[eb_id]    += hrs
                    emp_ot_days[eb_id].add(key)
            else:
                # Regular ('R' or unset): wh -> Shift cell, ot -> OT cell.
                if wh > 0:
                    code = shift_letter if wh == 8 else f"{shift_letter}1"
                    existing = emp_shifts[eb_id].get(key, "")
                    emp_shifts[eb_id][key] = (
                        f"{existing} {code}".strip() if existing else code
                    )
                    emp_whrs_by[eb_id][key] = round(
                        emp_whrs_by[eb_id].get(key, 0.0) + wh, 2
                    )
                    emp_whrs[eb_id]  += wh
                    emp_p_days[eb_id].add(key)
                if ot > 0:
                    code = shift_letter if ot == 8 else f"{shift_letter}1"
                    existing = emp_ot_shifts[eb_id].get(key, "")
                    emp_ot_shifts[eb_id][key] = (
                        f"{existing} {code}".strip() if existing else code
                    )
                    emp_ot_by[eb_id][key] = round(
                        emp_ot_by[eb_id].get(key, 0.0) + ot, 2
                    )
                    emp_ot[eb_id]    += ot
                    emp_ot_days[eb_id].add(key)

            wages = (rate / 8.0) * (wh + ot)
            emp_wages[eb_id][key] = round(
                emp_wages[eb_id].get(key, 0.0) + wages, 2
            )
            emp_total[eb_id] += wages
            if rate:
                emp_rate[eb_id] = rate

        out = []
        for eb_id, meta in emp_meta.items():
            out.append({
                **meta,
                "rate": round(emp_rate.get(eb_id, 0.0), 2),
                "shifts":         emp_shifts[eb_id],
                "ot_shifts":      emp_ot_shifts[eb_id],
                "wages":          emp_wages[eb_id],
                "working_hours":  emp_whrs_by[eb_id],
                "ot_hours":       emp_ot_by[eb_id],
                "total_working_hours": round(emp_whrs[eb_id], 2),
                "total_ot_hours":      round(emp_ot[eb_id],   2),
                "total_wages":         round(emp_total[eb_id], 2),
                "count_p_days":  len(emp_p_days[eb_id]),
                "count_ot_days": len(emp_ot_days[eb_id]),
            })
        # SQL already ORDER BY CAST(emp_code AS UNSIGNED), so dict-iteration order
        # of emp_meta preserves the desired numeric sort.
        return {"columns": columns, "data": out}
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")
